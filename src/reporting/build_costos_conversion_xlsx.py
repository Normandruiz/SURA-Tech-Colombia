"""
Análisis de costos de conversión - Google Ads + Meta Ads · Ene-Abr 2026

Lee:
  raw/google/2026_ene_abr/{Autos,Motos,Viajes,Arriendo}.csv  -- 1 fila por (campaña x mes)
  raw/meta/2026_ene_abr/Meta.csv                              -- 1 fila por (ad set x mes)

Produce:
  outputs/analisis_costos_conversion_ene_abr_2026.xlsx
    - Resumen ejecutivo (KPIs y comentarios)
    - Google Ads · mes a mes
    - Meta Ads · mes a mes
    - Combinado Google vs Meta
    - Variaciones MoM
"""

from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ----------------------------------------------------------------------------
# Paths
# ----------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parents[2]
GOOGLE_DIR = ROOT / "raw" / "google" / "2026_ene_abr"
META_PATH = ROOT / "raw" / "meta" / "2026_ene_abr" / "Meta.csv"
BI_PATH = ROOT / "docs" / "assets" / "data_bi.json"
OUT_PATH = ROOT / "outputs" / "analisis_costos_conversion_ene_abr_2026.xlsx"

# Map producto del script a sol del Sheet BI
SOL_MAP = {"Autos": "Autos", "Motos": "Motos", "Viajes": "Viajes", "Arriendo": "Arriendo"}
# Mes corto a key del Sheet BI
MES_TO_BI = {"Enero": "Enero", "Febrero": "Febrero", "Marzo": "Marzo", "Abril": "Abril"}
# Sources que mapean a cada plataforma
SOURCES_GOOGLE = {"google-ads"}
SOURCES_META   = {"meta-ads", "facebook", "ig", "instagram"}
SOURCES_BING   = {"bing-ads", "bing"}

PRODUCTS = ["Autos", "Motos", "Viajes", "Arriendo"]
MONTHS_ORDER = ["Enero", "Febrero", "Marzo", "Abril"]
MONTH_KEY = {"Enero": "Ene", "Febrero": "Feb", "Marzo": "Mar", "Abril": "Abr"}
MONTHS_SHORT = ["Ene", "Feb", "Mar", "Abr"]

# Files in raw/google/2026_ene_abr/
GOOGLE_FILES = {
    "Autos":    "autos.csv",
    "Motos":    "Motos.csv",
    "Viajes":   "viajes.csv",
    "Arriendo": "Arriendo.csv",
}

# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------
def parse_es_num(s: str) -> float:
    """ES-CO format: '1.234,56' -> 1234.56, '11.293' -> 11293.

    Reglas:
    - Si tiene coma: la coma es decimal, los puntos son separador miles.
    - Si solo tiene punto(s): si el patrón es 'X.YYY(.YYY)*' (grupos de 3 dígitos
      después de cada punto), se trata como separador de miles → se quita.
    - Si solo tiene un punto seguido de algo distinto a 3 dígitos, se trata
      como decimal (ej: '5.18').
    """
    if s is None:
        return 0.0
    s = str(s).strip().strip('"')
    if s in ("", "--", "-", "—"):
        return 0.0
    s = re.sub(r"[^\d,.\-]", "", s)
    if "," in s:
        # Coma es decimal en es-CO; puntos son thousands sep
        s = s.replace(".", "").replace(",", ".")
    else:
        # Sin coma. Verificar si los puntos son thousands separators.
        if re.match(r"^-?\d{1,3}(\.\d{3})+$", s):
            s = s.replace(".", "")
        # else: dejarlo (puede ser decimal como '5.18' o entero '1234')
    try:
        return float(s)
    except ValueError:
        return 0.0


def detect_product_meta(name: str) -> str:
    n = name.upper()
    # Order matters: ARRIENDO/ARRENDAMIENTO before others
    if "ARRIENDO" in n or "ARRENDAMIENTO" in n:
        return "Arriendo"
    if "VIAJE" in n:
        return "Viajes"
    if "MOTO" in n:
        return "Motos"
    if "AUTO" in n or "AUTOS" in n:
        return "Autos"
    return "Otro"


# ----------------------------------------------------------------------------
# Parsers
# ----------------------------------------------------------------------------
def parse_google_csv(path: Path) -> list[dict]:
    """Lee CSV Google Ads. Skip top 2 lines (header info), rest: campañas x mes."""
    rows = []
    with path.open("r", encoding="utf-8", newline="") as f:
        all_lines = f.readlines()
    # Skip first 2 lines (Informe de campaña + fecha range)
    csv_text = "".join(all_lines[2:])
    reader = csv.DictReader(csv_text.splitlines())
    for r in reader:
        mes_raw = (r.get("Mes") or "").strip()
        if not mes_raw:
            continue
        # "enero de 2026" -> "Enero"
        mes = mes_raw.split(" de ")[0].capitalize()
        if mes not in MONTHS_ORDER:
            continue
        # Filtrar filas de subtotal/total (Campaña vacía o con guiones)
        campana = (r.get("Campaña") or "").strip()
        if not campana or campana in ("--", "-", "—"):
            continue
        # Solo incluir campañas reales (que comiencen con SURA_ típicamente)
        if not campana.startswith("SURA"):
            continue
        rows.append({
            "mes":          mes,
            "campana":      (r.get("Campaña") or "").strip(),
            "estado":       (r.get("Estado de la campaña") or "").strip(),
            "tipo":         (r.get("Tipo de campaña") or "").strip(),
            "impresiones":  parse_es_num(r.get("Impr.") or ""),
            "clics":        parse_es_num(r.get("Clics") or ""),
            "ctr":          parse_es_num(r.get("Tasa de interacción") or ""),
            "coste":        parse_es_num(r.get("Coste") or ""),
            "cpc_medio":    parse_es_num(r.get("CPC medio") or ""),
            "conversiones": parse_es_num(r.get("Conversiones") or ""),
            "tasa_conv":    parse_es_num(r.get("Tasa de conv.") or ""),
            "cpa":          parse_es_num(r.get("Coste/conv.") or ""),
        })
    return rows


def parse_meta_csv(path: Path) -> list[dict]:
    rows = []
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            ini = (r.get("Inicio del informe") or "").strip()
            if not ini.startswith("2026-"):
                continue
            mes_num = ini.split("-")[1]
            mes = {"01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril"}.get(mes_num)
            if not mes:
                continue
            adset = (r.get("Nombre del conjunto de anuncios") or "").strip()
            prod = detect_product_meta(adset)
            spend_raw = r.get("Importe gastado (USD)") or "0"
            results_raw = r.get("Resultados") or "0"
            cpa_raw = r.get("Costo por resultados") or "0"
            impr_raw = r.get("Impresiones") or "0"
            try:
                spend = float(spend_raw) if spend_raw else 0.0
            except ValueError:
                spend = 0.0
            try:
                results = float(results_raw) if results_raw else 0.0
            except ValueError:
                results = 0.0
            try:
                cpa = float(cpa_raw) if cpa_raw else 0.0
            except ValueError:
                cpa = 0.0
            try:
                impr = int(float(impr_raw)) if impr_raw else 0
            except ValueError:
                impr = 0
            rows.append({
                "mes":          mes,
                "adset":        adset,
                "producto":     prod,
                "estado":       (r.get("Entrega del conjunto de anuncios") or "").strip(),
                "indicador":    (r.get("Indicador de resultado") or "").strip(),
                "spend":        spend,
                "results":      results,
                "cpa":          cpa,
                "impresiones":  impr,
            })
    return rows


# ----------------------------------------------------------------------------
# Aggregations
# ----------------------------------------------------------------------------
def aggregate_google(rows_by_product: dict[str, list[dict]]) -> dict:
    """ Devuelve { producto: { mes: {coste, conv, cpa, clics, impr} } } """
    out = {p: {m: {"coste": 0.0, "conv": 0.0, "clics": 0, "impr": 0}
               for m in MONTHS_ORDER}
           for p in PRODUCTS}
    for prod, rows in rows_by_product.items():
        for r in rows:
            agg = out[prod][r["mes"]]
            agg["coste"] += r["coste"]
            agg["conv"]  += r["conversiones"]
            agg["clics"] += int(r["clics"])
            agg["impr"]  += int(r["impresiones"])
    # CPA
    for p in PRODUCTS:
        for m in MONTHS_ORDER:
            a = out[p][m]
            a["cpa"] = a["coste"] / a["conv"] if a["conv"] > 0 else 0.0
    return out


def aggregate_meta(rows: list[dict]) -> dict:
    out = {p: {m: {"coste": 0.0, "conv": 0.0, "impr": 0}
               for m in MONTHS_ORDER}
           for p in PRODUCTS}
    for r in rows:
        if r["producto"] not in PRODUCTS:
            continue
        agg = out[r["producto"]][r["mes"]]
        agg["coste"] += r["spend"]
        agg["conv"]  += r["results"]
        agg["impr"]  += r["impresiones"]
    for p in PRODUCTS:
        for m in MONTHS_ORDER:
            a = out[p][m]
            a["cpa"] = a["coste"] / a["conv"] if a["conv"] > 0 else 0.0
            a["clics"] = 0  # Meta no expone clics directos en este export
    return out


# ----------------------------------------------------------------------------
# Excel styling
# ----------------------------------------------------------------------------
SURA_AZUL = "0033A0"
SURA_AQUA = "00AEC7"
SURA_GRIS_OSC = "53565A"
SURA_GRIS_CLARO = "F0F0F0"
VERDE = "059669"
ROJO = "DC2626"
AMARILLO = "F59E0B"

thin = Side(border_style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(cell):
    cell.fill = PatternFill("solid", fgColor=SURA_AZUL)
    cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border


def style_subheader(cell):
    cell.fill = PatternFill("solid", fgColor=SURA_AQUA)
    cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border


def style_total(cell):
    cell.fill = PatternFill("solid", fgColor=SURA_GRIS_CLARO)
    cell.font = Font(bold=True, color=SURA_GRIS_OSC, name="Calibri")
    cell.border = border


def style_data(cell):
    cell.font = Font(color=SURA_GRIS_OSC, name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="right")
    cell.border = border


# ----------------------------------------------------------------------------
# Sheets
# ----------------------------------------------------------------------------
def write_resumen(ws, google_agg, meta_agg):
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 26
    for col in "BCDEFGHIJKLMN":
        ws.column_dimensions[col].width = 14

    # Title
    ws["A1"] = "Análisis de costos de conversión · Ene–Abr 2026"
    ws["A1"].font = Font(bold=True, size=16, color=SURA_AZUL, name="Calibri")
    ws.merge_cells("A1:N1")

    ws["A2"] = "Google Ads + Meta Ads · 4 productos · conversiones reportadas por la plataforma"
    ws["A2"].font = Font(italic=True, color=SURA_GRIS_OSC, size=11)
    ws.merge_cells("A2:N2")

    row = 4
    # Headers row 1
    ws.cell(row=row, column=1).value = "Producto / Plataforma"
    style_header(ws.cell(row=row, column=1))
    col = 2
    for m in MONTHS_SHORT:
        for sub in ["Coste US$", "Conv.", "CPA US$"]:
            cell = ws.cell(row=row, column=col)
            cell.value = f"{m} · {sub}"
            style_header(cell)
            col += 1
    cell = ws.cell(row=row, column=col)
    cell.value = "Total Coste"
    style_header(cell)
    cell = ws.cell(row=row, column=col + 1)
    cell.value = "Total Conv."
    style_header(cell)
    ws.row_dimensions[row].height = 32
    row += 1

    # Data rows
    def write_block(label, agg):
        nonlocal row
        for prod in PRODUCTS:
            ws.cell(row=row, column=1).value = f"{prod} ({label})"
            ws.cell(row=row, column=1).font = Font(bold=True, color=SURA_AZUL, size=10)
            ws.cell(row=row, column=1).border = border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
            col = 2
            tot_coste, tot_conv = 0.0, 0.0
            for m in MONTHS_ORDER:
                a = agg[prod][m]
                ws.cell(row=row, column=col).value = round(a["coste"], 2)
                ws.cell(row=row, column=col + 1).value = round(a["conv"], 1)
                ws.cell(row=row, column=col + 2).value = round(a["cpa"], 2) if a["cpa"] else 0
                ws.cell(row=row, column=col).number_format = '"$"#,##0'
                ws.cell(row=row, column=col + 1).number_format = '#,##0'
                ws.cell(row=row, column=col + 2).number_format = '"$"#,##0.00'
                for k in range(3):
                    style_data(ws.cell(row=row, column=col + k))
                tot_coste += a["coste"]
                tot_conv += a["conv"]
                col += 3
            # Totals
            ws.cell(row=row, column=col).value = round(tot_coste, 2)
            ws.cell(row=row, column=col).number_format = '"$"#,##0'
            ws.cell(row=row, column=col + 1).value = round(tot_conv, 0)
            ws.cell(row=row, column=col + 1).number_format = '#,##0'
            for k in range(2):
                style_total(ws.cell(row=row, column=col + k))
            row += 1

    # Section header Google
    ws.cell(row=row, column=1).value = "GOOGLE ADS"
    style_subheader(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    row += 1
    write_block("Google", google_agg)

    # Section header Meta
    ws.cell(row=row, column=1).value = "META ADS"
    style_subheader(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    row += 1
    write_block("Meta", meta_agg)

    # Section combined
    ws.cell(row=row, column=1).value = "COMBINADO (Google + Meta)"
    style_subheader(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    row += 1
    combined = {p: {m: {
        "coste": google_agg[p][m]["coste"] + meta_agg[p][m]["coste"],
        "conv":  google_agg[p][m]["conv"]  + meta_agg[p][m]["conv"],
    } for m in MONTHS_ORDER} for p in PRODUCTS}
    for p in PRODUCTS:
        for m in MONTHS_ORDER:
            c = combined[p][m]
            c["cpa"] = c["coste"] / c["conv"] if c["conv"] > 0 else 0.0
    write_block("Total", combined)

    # Comentarios narrativos
    row += 2
    ws.cell(row=row, column=1).value = "INSIGHTS CLAVE"
    style_subheader(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    row += 1

    insights = build_insights(google_agg, meta_agg)
    for ins in insights:
        ws.cell(row=row, column=1).value = ins
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=1).font = Font(color=SURA_GRIS_OSC, size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
        ws.row_dimensions[row].height = 38
        row += 1


def build_insights(g: dict, m: dict) -> list[str]:
    """Genera bullets narrativos de evolución y CPA mes-a-mes por producto/plataforma."""
    out = []
    for prod in PRODUCTS:
        # Google insights
        ene_g = g[prod]["Enero"]
        abr_g = g[prod]["Abril"]
        if ene_g["conv"] > 0 and abr_g["conv"] > 0:
            d_conv = (abr_g["conv"] - ene_g["conv"]) / ene_g["conv"] * 100
            d_cpa  = (abr_g["cpa"]  - ene_g["cpa"])  / ene_g["cpa"]  * 100 if ene_g["cpa"] else 0
            sign_conv = "↑" if d_conv >= 0 else "↓"
            sign_cpa  = "↑" if d_cpa  >= 0 else "↓"
            out.append(
                f"• {prod} · Google Ads: conversiones {sign_conv} {abs(d_conv):.0f}% "
                f"(Ene {ene_g['conv']:,.0f} → Abr {abr_g['conv']:,.0f}) · "
                f"CPA {sign_cpa} {abs(d_cpa):.0f}% (US$ {ene_g['cpa']:.2f} → US$ {abr_g['cpa']:.2f}). "
                + ("Eficiencia degradada — revisar volumen Search vs PMax." if d_cpa > 15 else
                   "Eficiencia estable o mejorada." if d_cpa <= 5 else
                   "Leve degradación de eficiencia.")
            )
        # Meta insights
        ene_m = m[prod]["Enero"]; abr_m = m[prod]["Abril"]
        if ene_m["conv"] > 0 and abr_m["conv"] > 0:
            d_conv = (abr_m["conv"] - ene_m["conv"]) / ene_m["conv"] * 100
            d_cpa  = (abr_m["cpa"]  - ene_m["cpa"])  / ene_m["cpa"]  * 100 if ene_m["cpa"] else 0
            sign_conv = "↑" if d_conv >= 0 else "↓"
            sign_cpa  = "↑" if d_cpa  >= 0 else "↓"
            out.append(
                f"• {prod} · Meta Ads: conversiones {sign_conv} {abs(d_conv):.0f}% "
                f"(Ene {ene_m['conv']:,.0f} → Abr {abr_m['conv']:,.0f}) · "
                f"CPA {sign_cpa} {abs(d_cpa):.0f}% (US$ {ene_m['cpa']:.2f} → US$ {abr_m['cpa']:.2f}). "
                + ("Caída fuerte de volumen — fatiga creativa o saturación de audiencia." if d_conv < -25 else
                   "Volumen estable." if abs(d_conv) <= 10 else
                   "Volumen variable.")
            )
    return out


# ----------------------------------------------------------------------------
# Per-platform sheet
# ----------------------------------------------------------------------------
def write_platform(ws, title: str, agg: dict, color: str):
    ws.title = title
    ws.column_dimensions["A"].width = 18
    for col in "BCDEFGHIJKLMNOP":
        ws.column_dimensions[col].width = 13

    ws["A1"] = f"{title} · mes a mes · Ene–Abr 2026"
    ws["A1"].font = Font(bold=True, size=14, color=SURA_AZUL)
    ws.merge_cells("A1:Q1")

    # COSTE table
    row = 3
    ws.cell(row=row, column=1).value = "COSTE (US$)"
    style_subheader(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    headers = ["Producto"] + MONTHS_SHORT + ["Total"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i)
        c.value = h
        style_header(c)
    row += 1
    coste_table_start = row
    for prod in PRODUCTS:
        ws.cell(row=row, column=1).value = prod
        ws.cell(row=row, column=1).font = Font(bold=True, color=SURA_AZUL)
        ws.cell(row=row, column=1).border = border
        tot = 0.0
        for j, m in enumerate(MONTHS_ORDER):
            v = agg[prod][m]["coste"]
            c = ws.cell(row=row, column=2 + j)
            c.value = round(v, 2)
            c.number_format = '"$"#,##0'
            style_data(c)
            tot += v
        c = ws.cell(row=row, column=6)
        c.value = round(tot, 2)
        c.number_format = '"$"#,##0'
        style_total(c)
        row += 1

    # CONVERSIONES table
    row += 1
    ws.cell(row=row, column=1).value = "CONVERSIONES"
    style_subheader(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i)
        c.value = h
        style_header(c)
    row += 1
    conv_table_start = row
    for prod in PRODUCTS:
        ws.cell(row=row, column=1).value = prod
        ws.cell(row=row, column=1).font = Font(bold=True, color=SURA_AZUL)
        ws.cell(row=row, column=1).border = border
        tot = 0.0
        for j, m in enumerate(MONTHS_ORDER):
            v = agg[prod][m]["conv"]
            c = ws.cell(row=row, column=2 + j)
            c.value = round(v, 1)
            c.number_format = '#,##0'
            style_data(c)
            tot += v
        c = ws.cell(row=row, column=6)
        c.value = round(tot, 0)
        c.number_format = '#,##0'
        style_total(c)
        row += 1

    # CPA table
    row += 1
    ws.cell(row=row, column=1).value = "CPA (US$ por conversión)"
    style_subheader(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    cpa_header = ["Producto"] + MONTHS_SHORT + ["Promedio"]
    for i, h in enumerate(cpa_header, start=1):
        c = ws.cell(row=row, column=i)
        c.value = h
        style_header(c)
    row += 1
    cpa_table_start = row
    for prod in PRODUCTS:
        ws.cell(row=row, column=1).value = prod
        ws.cell(row=row, column=1).font = Font(bold=True, color=SURA_AZUL)
        ws.cell(row=row, column=1).border = border
        cpa_vals = []
        for j, m in enumerate(MONTHS_ORDER):
            v = agg[prod][m]["cpa"]
            c = ws.cell(row=row, column=2 + j)
            c.value = round(v, 2) if v else 0
            c.number_format = '"$"#,##0.00'
            style_data(c)
            if v > 0:
                cpa_vals.append(v)
        avg = sum(cpa_vals) / len(cpa_vals) if cpa_vals else 0
        c = ws.cell(row=row, column=6)
        c.value = round(avg, 2)
        c.number_format = '"$"#,##0.00'
        style_total(c)
        row += 1

    # MoM Variation table
    row += 1
    ws.cell(row=row, column=1).value = "VARIACIÓN MES-A-MES (CPA)"
    style_subheader(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    mom_header = ["Producto", "Ene→Feb", "Feb→Mar", "Mar→Abr", "Total Ene→Abr"]
    for i, h in enumerate(mom_header, start=1):
        c = ws.cell(row=row, column=i)
        c.value = h
        style_header(c)
    row += 1
    for prod in PRODUCTS:
        ws.cell(row=row, column=1).value = prod
        ws.cell(row=row, column=1).font = Font(bold=True, color=SURA_AZUL)
        ws.cell(row=row, column=1).border = border
        cpas = [agg[prod][m]["cpa"] for m in MONTHS_ORDER]
        deltas = []
        for i in range(1, 4):
            if cpas[i - 1] > 0:
                d = (cpas[i] - cpas[i - 1]) / cpas[i - 1] * 100
            else:
                d = 0
            deltas.append(d)
        # total
        if cpas[0] > 0 and cpas[3] > 0:
            d_tot = (cpas[3] - cpas[0]) / cpas[0] * 100
        else:
            d_tot = 0
        for j, d in enumerate(deltas + [d_tot]):
            c = ws.cell(row=row, column=2 + j)
            c.value = round(d, 1) / 100
            c.number_format = '0.0%;-0.0%'
            style_data(c)
            if d > 10:
                c.font = Font(bold=True, color=ROJO)
            elif d < -5:
                c.font = Font(bold=True, color=VERDE)
        row += 1

    # Charts
    # Chart CPA mes a mes (lineas)
    chart = LineChart()
    chart.title = "CPA mes a mes por producto (US$)"
    chart.style = 12
    chart.height = 10
    chart.width = 16
    data = Reference(ws, min_col=2, max_col=5, min_row=cpa_table_start - 1, max_row=cpa_table_start - 1 + len(PRODUCTS))
    cats = Reference(ws, min_col=1, max_col=1, min_row=cpa_table_start, max_row=cpa_table_start - 1 + len(PRODUCTS))
    # Note: we want products as series. Need to swap.
    # Actually our data layout: rows=products, cols=months. We want series=products, categories=months.
    chart2 = LineChart()
    chart2.title = f"{title}: CPA mensual por producto (US$)"
    chart2.style = 12
    chart2.height = 10
    chart2.width = 18
    chart2.y_axis.title = "CPA (US$)"
    chart2.x_axis.title = "Mes"
    data2 = Reference(ws, min_col=1, max_col=5, min_row=cpa_table_start, max_row=cpa_table_start - 1 + len(PRODUCTS))
    chart2.add_data(data2, titles_from_data=True, from_rows=True)
    cats2 = Reference(ws, min_col=2, max_col=5, min_row=cpa_table_start - 1, max_row=cpa_table_start - 1)
    chart2.set_categories(cats2)
    ws.add_chart(chart2, "H3")

    # Chart conversiones (barras agrupadas)
    chart3 = BarChart()
    chart3.type = "col"
    chart3.style = 11
    chart3.title = f"{title}: Conversiones mensuales por producto"
    chart3.y_axis.title = "Conversiones"
    chart3.x_axis.title = "Mes"
    chart3.height = 10
    chart3.width = 18
    data3 = Reference(ws, min_col=1, max_col=5, min_row=conv_table_start, max_row=conv_table_start - 1 + len(PRODUCTS))
    chart3.add_data(data3, titles_from_data=True, from_rows=True)
    cats3 = Reference(ws, min_col=2, max_col=5, min_row=conv_table_start - 1, max_row=conv_table_start - 1)
    chart3.set_categories(cats3)
    ws.add_chart(chart3, "H22")


import json


def load_bi_leads_by_product_month_source() -> dict:
    """ Devuelve {producto: {mes: {plataforma: leads}}} desde data_bi.json. """
    if not BI_PATH.exists():
        print(f"[WARN] missing BI sheet: {BI_PATH}")
        return {}
    data = json.loads(BI_PATH.read_text(encoding="utf-8"))
    out = {p: {m: {"google": 0, "meta": 0, "bing": 0, "otros": 0, "total": 0}
               for m in MONTHS_ORDER}
           for p in PRODUCTS}
    for mes, rows in data["months"].items():
        if mes not in MONTHS_ORDER:
            continue
        for r in rows:
            sol = r.get("sol")
            if sol not in PRODUCTS:
                continue
            src = (r.get("src") or "").strip().lower()
            leads = int(r.get("leads") or 0)
            if leads <= 0:
                continue
            target = out[sol][mes]
            target["total"] += leads
            if src in SOURCES_GOOGLE:
                target["google"] += leads
            elif src in SOURCES_META:
                target["meta"] += leads
            elif src in SOURCES_BING:
                target["bing"] += leads
            else:
                target["otros"] += leads
    return out


def write_cpl_negocio(ws, google_agg, meta_agg, bi_leads):
    """ Hoja: cruce Spend (CSV) vs Leads SF (Sheet BI) -> CPL real de negocio. """
    ws.title = "CPL Negocio"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    for col in "CDEFGHIJKLMN":
        ws.column_dimensions[col].width = 14

    ws["A1"] = "CPL de Negocio · Spend (CSV plataforma) ÷ Leads SF (Sheet BI / CRM)"
    ws["A1"].font = Font(bold=True, size=14, color=SURA_AZUL)
    ws.merge_cells("A1:N1")

    ws["A2"] = ("Conversiones plataforma (Google/Meta) NO siempre = lead real en SF. "
                "El CPL Negocio cruza inversión real con leads que llegaron al CRM por el canal correspondiente.")
    ws["A2"].font = Font(italic=True, color=SURA_GRIS_OSC, size=10)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:N2")
    ws.row_dimensions[2].height = 30

    row = 4
    headers = ["Producto", "Plataforma", "Mes",
               "Spend (US$)", "Conv. plataforma", "CPA plataforma",
               "Leads SF (CRM)", "CPL Negocio", "Gap CPL−CPA"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i)
        c.value = h
        style_header(c)
    ws.row_dimensions[row].height = 32
    row += 1
    table_start = row

    for prod in PRODUCTS:
        for plat, agg in [("Google", google_agg), ("Meta", meta_agg)]:
            for mes in MONTHS_ORDER:
                a = agg[prod][mes]
                spend = a["coste"]
                conv = a["conv"]
                cpa = a["cpa"]
                bi_key = "google" if plat == "Google" else "meta"
                leads_sf = bi_leads.get(prod, {}).get(mes, {}).get(bi_key, 0)
                cpl = (spend / leads_sf) if leads_sf > 0 else 0
                gap = (cpl - cpa) if (cpa > 0 and cpl > 0) else 0
                ws.cell(row=row, column=1).value = prod
                ws.cell(row=row, column=1).font = Font(bold=True, color=SURA_AZUL)
                ws.cell(row=row, column=1).border = border
                ws.cell(row=row, column=2).value = plat
                ws.cell(row=row, column=2).font = Font(italic=True, color=SURA_GRIS_OSC)
                ws.cell(row=row, column=2).border = border
                ws.cell(row=row, column=3).value = MONTH_KEY[mes]
                ws.cell(row=row, column=3).border = border
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=4).value = round(spend, 2)
                ws.cell(row=row, column=4).number_format = '"$"#,##0'
                style_data(ws.cell(row=row, column=4))
                ws.cell(row=row, column=5).value = round(conv, 0)
                ws.cell(row=row, column=5).number_format = '#,##0'
                style_data(ws.cell(row=row, column=5))
                ws.cell(row=row, column=6).value = round(cpa, 2) if cpa else 0
                ws.cell(row=row, column=6).number_format = '"$"#,##0.00'
                style_data(ws.cell(row=row, column=6))
                ws.cell(row=row, column=7).value = leads_sf
                ws.cell(row=row, column=7).number_format = '#,##0'
                style_data(ws.cell(row=row, column=7))
                ws.cell(row=row, column=8).value = round(cpl, 2) if cpl else 0
                ws.cell(row=row, column=8).number_format = '"$"#,##0.00'
                style_data(ws.cell(row=row, column=8))
                # Color gap
                ws.cell(row=row, column=9).value = round(gap, 2) if gap else 0
                ws.cell(row=row, column=9).number_format = '"$"#,##0.00'
                style_data(ws.cell(row=row, column=9))
                if gap > 5:
                    ws.cell(row=row, column=9).font = Font(bold=True, color=ROJO)
                elif gap < -2:
                    ws.cell(row=row, column=9).font = Font(bold=True, color=VERDE)
                row += 1

    # Tabla resumen Total Q1+Abril
    row += 2
    ws.cell(row=row, column=1).value = "RESUMEN TOTAL Ene–Abr (Inversión total / Leads SF / CPL Negocio)"
    style_subheader(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1
    res_headers = ["Producto", "Plataforma", "Inversión", "Leads SF", "CPL Negocio",
                   "Conv. plataforma", "CPA plataforma", "Δ CPL vs CPA", "% leads SF / conv plat."]
    for i, h in enumerate(res_headers, start=1):
        c = ws.cell(row=row, column=i)
        c.value = h
        style_header(c)
    ws.row_dimensions[row].height = 32
    row += 1
    for prod in PRODUCTS:
        for plat, agg in [("Google", google_agg), ("Meta", meta_agg)]:
            spend_tot = sum(agg[prod][m]["coste"] for m in MONTHS_ORDER)
            conv_tot = sum(agg[prod][m]["conv"] for m in MONTHS_ORDER)
            bi_key = "google" if plat == "Google" else "meta"
            leads_tot = sum(bi_leads.get(prod, {}).get(m, {}).get(bi_key, 0) for m in MONTHS_ORDER)
            cpl = spend_tot / leads_tot if leads_tot > 0 else 0
            cpa = spend_tot / conv_tot if conv_tot > 0 else 0
            gap = cpl - cpa if (cpa > 0 and cpl > 0) else 0
            ratio = leads_tot / conv_tot if conv_tot > 0 else 0
            ws.cell(row=row, column=1).value = prod
            ws.cell(row=row, column=1).font = Font(bold=True, color=SURA_AZUL)
            ws.cell(row=row, column=1).border = border
            ws.cell(row=row, column=2).value = plat
            ws.cell(row=row, column=2).font = Font(italic=True, color=SURA_GRIS_OSC)
            ws.cell(row=row, column=2).border = border
            ws.cell(row=row, column=3).value = round(spend_tot, 0)
            ws.cell(row=row, column=3).number_format = '"$"#,##0'
            style_total(ws.cell(row=row, column=3))
            ws.cell(row=row, column=4).value = leads_tot
            ws.cell(row=row, column=4).number_format = '#,##0'
            style_total(ws.cell(row=row, column=4))
            ws.cell(row=row, column=5).value = round(cpl, 2)
            ws.cell(row=row, column=5).number_format = '"$"#,##0.00'
            style_total(ws.cell(row=row, column=5))
            ws.cell(row=row, column=6).value = round(conv_tot, 0)
            ws.cell(row=row, column=6).number_format = '#,##0'
            style_data(ws.cell(row=row, column=6))
            ws.cell(row=row, column=7).value = round(cpa, 2)
            ws.cell(row=row, column=7).number_format = '"$"#,##0.00'
            style_data(ws.cell(row=row, column=7))
            ws.cell(row=row, column=8).value = round(gap, 2)
            ws.cell(row=row, column=8).number_format = '"$"#,##0.00'
            style_data(ws.cell(row=row, column=8))
            if gap > 5:
                ws.cell(row=row, column=8).font = Font(bold=True, color=ROJO)
            elif gap < -2:
                ws.cell(row=row, column=8).font = Font(bold=True, color=VERDE)
            ws.cell(row=row, column=9).value = ratio
            ws.cell(row=row, column=9).number_format = '0%'
            style_data(ws.cell(row=row, column=9))
            row += 1

    # Insights narrativos sobre CPL vs CPA
    row += 2
    ws.cell(row=row, column=1).value = "INSIGHTS · CPL Negocio vs CPA Plataforma"
    style_subheader(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1
    insights = []
    for prod in PRODUCTS:
        for plat, agg in [("Google", google_agg), ("Meta", meta_agg)]:
            spend_tot = sum(agg[prod][m]["coste"] for m in MONTHS_ORDER)
            conv_tot = sum(agg[prod][m]["conv"] for m in MONTHS_ORDER)
            bi_key = "google" if plat == "Google" else "meta"
            leads_tot = sum(bi_leads.get(prod, {}).get(m, {}).get(bi_key, 0) for m in MONTHS_ORDER)
            if leads_tot == 0 or conv_tot == 0:
                continue
            cpl = spend_tot / leads_tot
            cpa = spend_tot / conv_tot
            ratio = leads_tot / conv_tot
            if ratio > 1:
                txt = (f"• {prod} · {plat}: CRM reportó MÁS leads ({leads_tot:,.0f}) que conversiones plataforma ({conv_tot:,.0f}). "
                       f"CPL Negocio US$ {cpl:.2f} vs CPA US$ {cpa:.2f}. Tracking robusto, CPL real es MENOR.")
            elif ratio < 0.3:
                txt = (f"• {prod} · {plat}: ratio leads/conv muy bajo ({ratio*100:.0f}%) — "
                       f"plataforma reporta {conv_tot:,.0f} pero CRM solo recibió {leads_tot:,.0f}. "
                       f"CPL Negocio US$ {cpl:.2f} vs CPA US$ {cpa:.2f}. Posible micro-conversión inflada.")
            else:
                txt = (f"• {prod} · {plat}: CPL Negocio US$ {cpl:.2f}, CPA US$ {cpa:.2f}, "
                       f"ratio leads/conv {ratio*100:.0f}%.")
            insights.append(txt)
    for ins in insights:
        ws.cell(row=row, column=1).value = ins
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=1).font = Font(color=SURA_GRIS_OSC, size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        ws.row_dimensions[row].height = 36
        row += 1


def load_bi_full() -> dict:
    """ Carga todo el Sheet BI: rows por mes con sol/src/camp/cont/leads/pol/prima. """
    if not BI_PATH.exists():
        return {}
    return json.loads(BI_PATH.read_text(encoding="utf-8"))


def index_bi_by_campaign(bi_data: dict) -> dict:
    """Index by (sol_norm, camp_norm) -> {mes: {leads, pol, prima}}.

    Match heuristic: case-insensitive, igualar guion bajo y guion medio.
    """
    out = defaultdict(lambda: defaultdict(lambda: {"leads": 0, "pol": 0, "prima": 0}))
    for mes, rows in bi_data.get("months", {}).items():
        if mes not in MONTHS_ORDER:
            continue
        for r in rows:
            sol = r.get("sol")
            camp = (r.get("camp") or "").strip()
            if not sol or not camp:
                continue
            key = (sol, camp.lower().strip())
            agg = out[key][mes]
            agg["leads"] += int(r.get("leads") or 0)
            agg["pol"]   += int(r.get("pol") or 0)
            agg["prima"] += int(r.get("prima") or 0)
    return out


def match_campaign_to_bi(producto: str, campaign_name: str, bi_index: dict) -> dict:
    """Devuelve {mes: {leads, pol, prima}} para la campaña matcheada.

    Match exacto case-insensitive. Si no matchea, devuelve dict vacío.
    """
    target = campaign_name.lower().strip()
    key = (producto, target)
    if key in bi_index:
        return dict(bi_index[key])
    # Fallback: buscar match parcial (campaign name contiene la key del Sheet BI o viceversa)
    for (sol, camp_key), data in bi_index.items():
        if sol != producto:
            continue
        if camp_key == target or camp_key.replace("_","-") == target.replace("_","-"):
            return dict(data)
    return {}


def write_match_campanas(ws, google_rows: dict[str, list[dict]], meta_rows: list[dict],
                         bi_index: dict):
    """ Hoja: cada campaña/ad set vs leads SF matcheados. """
    ws.title = "Match Campañas"
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 50
    for col in "DEFGHIJK":
        ws.column_dimensions[col].width = 13

    ws["A1"] = "Match campaña-a-campaña · Inversión vs Leads SF (CRM) por campaña individual"
    ws["A1"].font = Font(bold=True, size=14, color=SURA_AZUL)
    ws.merge_cells("A1:K1")

    ws["A2"] = ("Cada fila = una campaña Google o ad set Meta (totales Ene-Abr). "
                "Match: nombre campaña/adset → 'Campaign UTM' del Sheet BI (case-insensitive). "
                "Filas marcadas en ROJO = inversión > 0 pero leads SF = 0 (ROI cero).")
    ws["A2"].font = Font(italic=True, color=SURA_GRIS_OSC, size=10)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:K2")
    ws.row_dimensions[2].height = 36

    row = 4
    headers = ["Producto", "Plataforma", "Campaña", "Estado",
               "Spend (US$)", "Conv. Plat.", "CPA Plat.",
               "Leads SF", "CPL Negocio", "Pólizas", "CAC"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i)
        c.value = h
        style_header(c)
    ws.row_dimensions[row].height = 32
    row += 1

    # Aggregate Google: por (producto, campaña) -> totales Ene-Abr
    google_by_camp = defaultdict(lambda: {"spend": 0, "conv": 0, "estado": ""})
    for prod, rows in google_rows.items():
        for r in rows:
            k = (prod, r["campana"])
            google_by_camp[k]["spend"] += r["coste"]
            google_by_camp[k]["conv"]  += r["conversiones"]
            google_by_camp[k]["estado"] = r["estado"] or google_by_camp[k]["estado"]

    # Aggregate Meta: por (producto, adset) -> totales Ene-Abr
    meta_by_adset = defaultdict(lambda: {"spend": 0, "conv": 0, "estado": ""})
    for r in meta_rows:
        if r["producto"] not in PRODUCTS:
            continue
        k = (r["producto"], r["adset"])
        meta_by_adset[k]["spend"] += r["spend"]
        meta_by_adset[k]["conv"]  += r["results"]
        meta_by_adset[k]["estado"] = r["estado"] or meta_by_adset[k]["estado"]

    # Combine and sort by spend desc
    all_items = []
    for (prod, camp), v in google_by_camp.items():
        all_items.append((prod, "Google", camp, v["estado"], v["spend"], v["conv"]))
    for (prod, adset), v in meta_by_adset.items():
        all_items.append((prod, "Meta", adset, v["estado"], v["spend"], v["conv"]))
    all_items.sort(key=lambda x: -x[4])  # by spend desc

    roi_zero_count = 0
    for prod, plat, camp, estado, spend, conv in all_items:
        # Match con Sheet BI
        matched = match_campaign_to_bi(prod, camp, bi_index)
        leads_sf = sum(d["leads"] for d in matched.values())
        polizas  = sum(d["pol"]   for d in matched.values())
        cpl = (spend / leads_sf) if leads_sf > 0 else 0
        cac = (spend / polizas)  if polizas > 0  else 0
        cpa = (spend / conv)     if conv > 0     else 0

        ws.cell(row=row, column=1).value = prod
        ws.cell(row=row, column=1).font = Font(bold=True, color=SURA_AZUL, size=10)
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2).value = plat
        ws.cell(row=row, column=2).font = Font(italic=True, color=SURA_GRIS_OSC, size=10)
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=3).value = camp
        ws.cell(row=row, column=3).font = Font(color=SURA_GRIS_OSC, size=9)
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=4).value = estado or "—"
        ws.cell(row=row, column=4).font = Font(color=SURA_GRIS_OSC, size=9)
        ws.cell(row=row, column=4).border = border
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5).value = round(spend, 0)
        ws.cell(row=row, column=5).number_format = '"$"#,##0'
        style_data(ws.cell(row=row, column=5))
        ws.cell(row=row, column=6).value = round(conv, 0)
        ws.cell(row=row, column=6).number_format = '#,##0'
        style_data(ws.cell(row=row, column=6))
        ws.cell(row=row, column=7).value = round(cpa, 2) if cpa else 0
        ws.cell(row=row, column=7).number_format = '"$"#,##0.00'
        style_data(ws.cell(row=row, column=7))
        ws.cell(row=row, column=8).value = leads_sf
        ws.cell(row=row, column=8).number_format = '#,##0'
        style_data(ws.cell(row=row, column=8))
        ws.cell(row=row, column=9).value = round(cpl, 2) if cpl else 0
        ws.cell(row=row, column=9).number_format = '"$"#,##0.00'
        style_data(ws.cell(row=row, column=9))
        ws.cell(row=row, column=10).value = polizas
        ws.cell(row=row, column=10).number_format = '#,##0'
        style_data(ws.cell(row=row, column=10))
        ws.cell(row=row, column=11).value = round(cac, 0) if cac else 0
        ws.cell(row=row, column=11).number_format = '"$"#,##0'
        style_data(ws.cell(row=row, column=11))

        # ROI cero: spend > 50 USD pero leads_sf == 0
        if spend > 50 and leads_sf == 0:
            roi_zero_count += 1
            for cidx in range(1, 12):
                cell = ws.cell(row=row, column=cidx)
                cell.fill = PatternFill("solid", fgColor="FEE2E2")
            ws.cell(row=row, column=8).font = Font(bold=True, color=ROJO)
        elif leads_sf > 0 and cpl > 0 and cpa > 0 and cpl < cpa * 0.5:
            # CPL real menor a la mitad de CPA plataforma = oportunidad
            ws.cell(row=row, column=9).font = Font(bold=True, color=VERDE)
        row += 1

    # Bottom: contador
    row += 1
    ws.cell(row=row, column=1).value = (
        f"Campañas con ROI cero (Spend > $50 USD, Leads SF = 0): {roi_zero_count}. "
        "Revisar UTMs / atribución / pausar."
    )
    ws.cell(row=row, column=1).font = Font(italic=True, color=ROJO, bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)


def write_cac(ws, google_agg, meta_agg, bi_data):
    """ Hoja: CAC real (Spend / Pólizas del Sheet BI) por producto x plataforma x mes. """
    ws.title = "CAC Real"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    for col in "CDEFGHIJK":
        ws.column_dimensions[col].width = 14

    ws["A1"] = "CAC Real · Inversión ÷ Pólizas Emitidas (Sheet BI)"
    ws["A1"].font = Font(bold=True, size=14, color=SURA_AZUL)
    ws.merge_cells("A1:K1")

    ws["A2"] = ("CAC = costo de adquirir 1 venta cerrada (póliza emitida). "
                "Es el indicador FINAL del negocio — no leads, no clicks: pólizas reales que pagan prima.")
    ws["A2"].font = Font(italic=True, color=SURA_GRIS_OSC, size=10)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:K2")
    ws.row_dimensions[2].height = 30

    # Compute polizas by (producto, mes, plataforma) from Sheet BI
    pol_by_pms = {p: {m: {"google": 0, "meta": 0} for m in MONTHS_ORDER} for p in PRODUCTS}
    prima_by_pms = {p: {m: {"google": 0, "meta": 0} for m in MONTHS_ORDER} for p in PRODUCTS}
    for mes, rows in bi_data.get("months", {}).items():
        if mes not in MONTHS_ORDER:
            continue
        for r in rows:
            sol = r.get("sol")
            if sol not in PRODUCTS:
                continue
            src = (r.get("src") or "").lower()
            pol = int(r.get("pol") or 0)
            prima = int(r.get("prima") or 0)
            if src in SOURCES_GOOGLE:
                pol_by_pms[sol][mes]["google"] += pol
                prima_by_pms[sol][mes]["google"] += prima
            elif src in SOURCES_META:
                pol_by_pms[sol][mes]["meta"] += pol
                prima_by_pms[sol][mes]["meta"] += prima

    row = 4
    headers = ["Producto", "Plataforma", "Mes",
               "Spend (US$)", "Pólizas", "CAC (US$)",
               "Prima emitida (COP)", "ROAS (Prima/Spend)*"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i)
        c.value = h
        style_header(c)
    ws.row_dimensions[row].height = 32
    row += 1

    # COP/USD assumption (informativo): 4000 COP = 1 USD
    COP_USD = 4000
    for prod in PRODUCTS:
        for plat, agg in [("Google", google_agg), ("Meta", meta_agg)]:
            for mes in MONTHS_ORDER:
                spend = agg[prod][mes]["coste"]
                bi_key = "google" if plat == "Google" else "meta"
                pol = pol_by_pms[prod][mes][bi_key]
                prima_cop = prima_by_pms[prod][mes][bi_key]
                cac = spend / pol if pol > 0 else 0
                prima_usd = prima_cop / COP_USD
                roas = prima_usd / spend if spend > 0 else 0
                ws.cell(row=row, column=1).value = prod
                ws.cell(row=row, column=1).font = Font(bold=True, color=SURA_AZUL)
                ws.cell(row=row, column=1).border = border
                ws.cell(row=row, column=2).value = plat
                ws.cell(row=row, column=2).font = Font(italic=True, color=SURA_GRIS_OSC)
                ws.cell(row=row, column=2).border = border
                ws.cell(row=row, column=3).value = MONTH_KEY[mes]
                ws.cell(row=row, column=3).border = border
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=4).value = round(spend, 0)
                ws.cell(row=row, column=4).number_format = '"$"#,##0'
                style_data(ws.cell(row=row, column=4))
                ws.cell(row=row, column=5).value = pol
                ws.cell(row=row, column=5).number_format = '#,##0'
                style_data(ws.cell(row=row, column=5))
                ws.cell(row=row, column=6).value = round(cac, 0) if cac else 0
                ws.cell(row=row, column=6).number_format = '"$"#,##0'
                style_data(ws.cell(row=row, column=6))
                ws.cell(row=row, column=7).value = prima_cop
                ws.cell(row=row, column=7).number_format = '"$"#,##0'
                style_data(ws.cell(row=row, column=7))
                ws.cell(row=row, column=8).value = round(roas, 1) if roas else 0
                ws.cell(row=row, column=8).number_format = '0.0"x"'
                style_data(ws.cell(row=row, column=8))
                if roas >= 5:
                    ws.cell(row=row, column=8).font = Font(bold=True, color=VERDE)
                elif roas > 0 and roas < 1:
                    ws.cell(row=row, column=8).font = Font(bold=True, color=ROJO)
                row += 1

    # Resumen total
    row += 2
    ws.cell(row=row, column=1).value = "RESUMEN TOTAL Ene–Abr · CAC y ROAS"
    style_subheader(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1
    res_headers = ["Producto", "Plataforma", "Inversión", "Pólizas", "CAC", "Prima emitida (COP)", "Prima (US$)*", "ROAS"]
    for i, h in enumerate(res_headers, start=1):
        c = ws.cell(row=row, column=i)
        c.value = h
        style_header(c)
    row += 1
    for prod in PRODUCTS:
        for plat, agg in [("Google", google_agg), ("Meta", meta_agg)]:
            spend_tot = sum(agg[prod][m]["coste"] for m in MONTHS_ORDER)
            bi_key = "google" if plat == "Google" else "meta"
            pol_tot = sum(pol_by_pms[prod][m][bi_key] for m in MONTHS_ORDER)
            prima_cop = sum(prima_by_pms[prod][m][bi_key] for m in MONTHS_ORDER)
            cac = spend_tot / pol_tot if pol_tot > 0 else 0
            prima_usd = prima_cop / COP_USD
            roas = prima_usd / spend_tot if spend_tot > 0 else 0
            ws.cell(row=row, column=1).value = prod
            ws.cell(row=row, column=1).font = Font(bold=True, color=SURA_AZUL)
            ws.cell(row=row, column=1).border = border
            ws.cell(row=row, column=2).value = plat
            ws.cell(row=row, column=2).font = Font(italic=True, color=SURA_GRIS_OSC)
            ws.cell(row=row, column=2).border = border
            ws.cell(row=row, column=3).value = round(spend_tot, 0)
            ws.cell(row=row, column=3).number_format = '"$"#,##0'
            style_total(ws.cell(row=row, column=3))
            ws.cell(row=row, column=4).value = pol_tot
            ws.cell(row=row, column=4).number_format = '#,##0'
            style_total(ws.cell(row=row, column=4))
            ws.cell(row=row, column=5).value = round(cac, 0) if cac else 0
            ws.cell(row=row, column=5).number_format = '"$"#,##0'
            style_total(ws.cell(row=row, column=5))
            ws.cell(row=row, column=6).value = prima_cop
            ws.cell(row=row, column=6).number_format = '"$"#,##0'
            style_data(ws.cell(row=row, column=6))
            ws.cell(row=row, column=7).value = round(prima_usd, 0)
            ws.cell(row=row, column=7).number_format = '"$"#,##0'
            style_data(ws.cell(row=row, column=7))
            ws.cell(row=row, column=8).value = round(roas, 1) if roas else 0
            ws.cell(row=row, column=8).number_format = '0.0"x"'
            style_total(ws.cell(row=row, column=8))
            if roas >= 5:
                ws.cell(row=row, column=8).font = Font(bold=True, color=VERDE, size=11)
            elif roas > 0 and roas < 1:
                ws.cell(row=row, column=8).font = Font(bold=True, color=ROJO, size=11)
            row += 1

    row += 1
    ws.cell(row=row, column=1).value = "* ROAS calculado con TC referencial 4.000 COP/USD"
    ws.cell(row=row, column=1).font = Font(italic=True, color=SURA_GRIS_OSC, size=9)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)


def aggregate_google_full(rows_by_product: dict[str, list[dict]]) -> dict:
    """ Como aggregate_google pero conservando clics e impresiones para descomposición."""
    out = {p: {m: {"coste": 0.0, "conv": 0.0, "clics": 0, "impr": 0}
               for m in MONTHS_ORDER}
           for p in PRODUCTS}
    for prod, rows in rows_by_product.items():
        for r in rows:
            agg = out[prod][r["mes"]]
            agg["coste"] += r["coste"]
            agg["conv"]  += r["conversiones"]
            agg["clics"] += int(r["clics"])
            agg["impr"]  += int(r["impresiones"])
    for p in PRODUCTS:
        for m in MONTHS_ORDER:
            a = out[p][m]
            a["cpc"] = a["coste"] / a["clics"] if a["clics"] > 0 else 0
            a["ctr"] = a["clics"] / a["impr"] if a["impr"] > 0 else 0
            a["tasa_conv"] = a["conv"] / a["clics"] if a["clics"] > 0 else 0
            a["cpa"] = a["coste"] / a["conv"] if a["conv"] > 0 else 0
            a["cpm"] = (a["coste"] * 1000) / a["impr"] if a["impr"] > 0 else 0
    return out


def aggregate_meta_full(rows: list[dict]) -> dict:
    """Meta no expone clicks pero sí impresiones."""
    out = {p: {m: {"coste": 0.0, "conv": 0.0, "impr": 0}
               for m in MONTHS_ORDER}
           for p in PRODUCTS}
    for r in rows:
        if r["producto"] not in PRODUCTS:
            continue
        agg = out[r["producto"]][r["mes"]]
        agg["coste"] += r["spend"]
        agg["conv"]  += r["results"]
        agg["impr"]  += r["impresiones"]
    for p in PRODUCTS:
        for m in MONTHS_ORDER:
            a = out[p][m]
            a["cpa"] = a["coste"] / a["conv"] if a["conv"] > 0 else 0
            a["cpm"] = (a["coste"] * 1000) / a["impr"] if a["impr"] > 0 else 0
            a["clics"] = 0
            a["cpc"] = 0
            a["ctr"] = 0
            a["tasa_conv"] = a["conv"] / a["impr"] if a["impr"] > 0 else 0
    return out


def diagnose_cpl_change(ene: dict, abr: dict, plat: str) -> dict:
    """Devuelve un diagnóstico estructurado del cambio de CPL Ene→Abr."""
    def pct(a, b):
        return ((b - a) / a * 100) if a > 0 else 0

    d = {
        "spend":   pct(ene["coste"], abr["coste"]),
        "conv":    pct(ene["conv"],  abr["conv"]),
        "impr":    pct(ene["impr"],  abr["impr"]),
        "cpa":     pct(ene["cpa"],   abr["cpa"]),
        "cpm":     pct(ene["cpm"],   abr["cpm"]),
    }
    if plat == "Google":
        d["clics"]      = pct(ene["clics"],     abr["clics"])
        d["cpc"]        = pct(ene["cpc"],       abr["cpc"])
        d["ctr"]        = pct(ene["ctr"],       abr["ctr"])
        d["tasa_conv"]  = pct(ene["tasa_conv"], abr["tasa_conv"])
    else:
        # Meta: tasa_conv es conv/impr
        d["tasa_conv"]  = pct(ene["tasa_conv"], abr["tasa_conv"])

    # Hipótesis de causa raíz
    drivers = []
    actions = []
    severity = "OK"

    if d["cpa"] > 20:
        severity = "CRÍTICO" if d["cpa"] > 80 else "ALTO" if d["cpa"] > 40 else "MEDIO"

    if plat == "Google":
        # Diagnóstico Google
        if d.get("cpc", 0) > 20 and d.get("tasa_conv", 0) > -10:
            drivers.append(f"Subida fuerte de CPC (+{d['cpc']:.0f}%) — competencia o auction más cara")
            actions.append("Revisar Auction Insights, ajustar pujas, validar Quality Score")
        if d.get("tasa_conv", 0) < -20 and d.get("cpc", 0) < 15:
            drivers.append(f"Caída de tasa de conversión ({d['tasa_conv']:+.0f}%) — calidad de tráfico o landing")
            actions.append("Auditar landing pages, revisar Core Web Vitals, validar formularios")
        if d.get("cpc", 0) > 20 and d.get("tasa_conv", 0) < -15:
            drivers.append(f"DOBLE FACTOR: CPC sube (+{d['cpc']:.0f}%) Y tasa conv cae ({d['tasa_conv']:+.0f}%)")
            actions.append("Auditar full funnel: bidding strategy + landing experience")
        if d["impr"] < -20:
            drivers.append(f"Impresiones cayeron {d['impr']:+.0f}% — pérdida de cobertura del mercado")
            actions.append("Revisar Lost IS Budget y Lost IS Rank en panel Cuotas")
        if d.get("ctr", 0) < -15:
            drivers.append(f"CTR cae ({d['ctr']:+.0f}%) — fatiga creativa o RSAs poco optimizadas")
            actions.append("Refrescar headlines/descriptions, A/B test de copy nuevo")
    else:
        # Diagnóstico Meta
        if d["cpm"] > 20:
            drivers.append(f"CPM sube +{d['cpm']:.0f}% — saturación audiencia o competencia interna")
            actions.append("Ampliar audiencias o reducir frecuencia / lookalikes nuevos")
        if d.get("tasa_conv", 0) < -20:
            drivers.append(f"Tasa conv (conv/impr) cae {d['tasa_conv']:+.0f}% — fatiga creativa")
            actions.append("Refrescar creativos: imagen → video, nuevos hooks, UGC")
        if d["impr"] < -20 and d["spend"] < -10:
            drivers.append(f"Recorte de inversión Meta ({d['spend']:+.0f}%) — pero CPA igual subió")
            actions.append("Auditar audiencias activas: pueden estar saturadas pese a menor budget")

    if not drivers:
        if d["cpa"] <= 5:
            drivers.append("CPL estable — performance consistente")
        elif d["cpa"] > 5:
            drivers.append("Aumento moderado de CPL sin causa única identificable")
            actions.append("Revisión integral mes-a-mes en próximo periodo")

    d["drivers"] = drivers
    d["actions"] = actions
    d["severity"] = severity
    return d


def write_diagnostico_ceo(ws, google_full, meta_full):
    """ Hoja: análisis de drivers de aumento CPL para presentar al CEO. """
    ws.title = "Diagnóstico CEO"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 9
    for col in "CDEFGHIJK":
        ws.column_dimensions[col].width = 10

    ws["A1"] = "Diagnóstico forense del aumento de CPL · Ene vs Abr 2026"
    ws["A1"].font = Font(bold=True, size=14, color=SURA_AZUL)
    ws.merge_cells("A1:K1")

    ws["A2"] = ("Por producto y plataforma: descomposición del cambio de CPL en sus drivers (CPC, CPM, CTR, "
                "Tasa Conv, Impresiones), con hipótesis de causa raíz y acción sugerida. "
                "Objetivo: dar al CEO un diagnóstico estructurado, no solo 'el CPL subió'.")
    ws["A2"].font = Font(italic=True, color=SURA_GRIS_OSC, size=10)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:K2")
    ws.row_dimensions[2].height = 36

    row = 4
    # Tabla descomposición
    ws.cell(row=row, column=1).value = "Δ% Ene→Abr 2026 por driver (factores que mueven el CPL)"
    style_subheader(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    row += 1
    headers = ["Producto", "Plat.", "Severidad CPL",
               "ΔCPL", "ΔSpend", "ΔConv", "ΔImpr",
               "ΔCPC", "ΔCTR", "ΔTasaConv", "Drivers principales"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i)
        c.value = h
        style_header(c)
    ws.row_dimensions[row].height = 36
    row += 1

    diagnoses = []
    for prod in PRODUCTS:
        for plat, agg in [("Google", google_full), ("Meta", meta_full)]:
            ene = agg[prod]["Enero"]
            abr = agg[prod]["Abril"]
            d = diagnose_cpl_change(ene, abr, plat)
            diagnoses.append((prod, plat, d, ene, abr))

            ws.cell(row=row, column=1).value = prod
            ws.cell(row=row, column=1).font = Font(bold=True, color=SURA_AZUL, size=10)
            ws.cell(row=row, column=1).border = border
            ws.cell(row=row, column=2).value = plat
            ws.cell(row=row, column=2).font = Font(italic=True, color=SURA_GRIS_OSC, size=10)
            ws.cell(row=row, column=2).border = border

            sev = d["severity"]
            sev_color = {"CRÍTICO": ROJO, "ALTO": "EA580C", "MEDIO": AMARILLO, "OK": VERDE}.get(sev, SURA_GRIS_OSC)
            ws.cell(row=row, column=3).value = sev
            ws.cell(row=row, column=3).font = Font(bold=True, color=sev_color, size=10)
            ws.cell(row=row, column=3).border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

            def write_pct(col, val, threshold_bad=20, threshold_good=-10):
                c = ws.cell(row=row, column=col)
                c.value = val / 100
                c.number_format = '0.0%;-0.0%;0.0%'
                style_data(c)
                if val > threshold_bad:
                    c.font = Font(bold=True, color=ROJO, size=10)
                elif val < threshold_good:
                    c.font = Font(bold=True, color=VERDE, size=10)

            write_pct(4, d["cpa"])           # ΔCPL
            write_pct(5, d["spend"], 30, -10)
            write_pct(6, d["conv"], 0, -20)  # baja de conv = malo
            write_pct(7, d["impr"], 0, -20)
            if plat == "Google":
                write_pct(8, d.get("cpc", 0))
                write_pct(9, d.get("ctr", 0), 0, -15)
                write_pct(10, d.get("tasa_conv", 0), 0, -15)
            else:
                ws.cell(row=row, column=8).value = "n/a"
                ws.cell(row=row, column=8).border = border
                ws.cell(row=row, column=8).font = Font(color=SURA_GRIS_OSC, italic=True, size=9)
                ws.cell(row=row, column=9).value = "n/a"
                ws.cell(row=row, column=9).border = border
                ws.cell(row=row, column=9).font = Font(color=SURA_GRIS_OSC, italic=True, size=9)
                write_pct(10, d.get("tasa_conv", 0), 0, -15)

            drv_text = " · ".join(d["drivers"][:2]) if d["drivers"] else "—"
            ws.cell(row=row, column=11).value = drv_text
            ws.cell(row=row, column=11).font = Font(color=SURA_GRIS_OSC, size=9)
            ws.cell(row=row, column=11).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row, column=11).border = border
            ws.column_dimensions[get_column_letter(11)].width = 50
            ws.row_dimensions[row].height = 42
            row += 1

    # Sección narrativa por producto: "Caso CEO"
    row += 2
    ws.cell(row=row, column=1).value = "Casos para CEO · narrativa diagnóstica por producto-plataforma"
    style_subheader(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    row += 2

    # Solo casos con CPL en aumento de severidad MEDIO o más
    cases = [(p, pl, d, e, a) for p, pl, d, e, a in diagnoses if d["severity"] in ("MEDIO", "ALTO", "CRÍTICO")]
    cases.sort(key=lambda x: -x[2]["cpa"])  # por mayor aumento CPL primero

    for prod, plat, d, ene, abr in cases:
        # Title
        ws.cell(row=row, column=1).value = f"{prod} · {plat}  —  CPL Ene US$ {ene['cpa']:.2f} → Abr US$ {abr['cpa']:.2f}  ({d['cpa']:+.0f}%)"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color=SURA_AZUL)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        row += 1

        # Severity tag
        sev = d["severity"]
        sev_label = {"CRÍTICO": "🚨 CRÍTICO", "ALTO": "⚠️ ALTO", "MEDIO": "🟡 MEDIO", "OK": "✓ OK"}.get(sev, sev)
        sev_color = {"CRÍTICO": ROJO, "ALTO": "EA580C", "MEDIO": AMARILLO, "OK": VERDE}.get(sev, SURA_GRIS_OSC)
        ws.cell(row=row, column=1).value = f"Severidad: {sev_label}"
        ws.cell(row=row, column=1).font = Font(bold=True, color=sev_color, size=11)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        row += 1

        # Datos clave
        spend_ene = ene["coste"]; spend_abr = abr["coste"]
        conv_ene = ene["conv"]; conv_abr = abr["conv"]
        ws.cell(row=row, column=1).value = (
            f"Datos: Inversión {spend_ene:,.0f} → {spend_abr:,.0f} USD ({d['spend']:+.0f}%) · "
            f"Conversiones {conv_ene:,.0f} → {conv_abr:,.0f} ({d['conv']:+.0f}%)"
        )
        ws.cell(row=row, column=1).font = Font(color=SURA_GRIS_OSC, size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        row += 1

        # Drivers (causas)
        ws.cell(row=row, column=1).value = "🔬 Causas raíz identificadas:"
        ws.cell(row=row, column=1).font = Font(bold=True, color=SURA_AZUL, size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        row += 1
        for drv in d["drivers"]:
            ws.cell(row=row, column=1).value = f"   • {drv}"
            ws.cell(row=row, column=1).font = Font(color=SURA_GRIS_OSC, size=10)
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
            ws.row_dimensions[row].height = 26
            row += 1

        # Actions (qué hacer)
        if d["actions"]:
            ws.cell(row=row, column=1).value = "🎯 Plan de acción propuesto:"
            ws.cell(row=row, column=1).font = Font(bold=True, color=SURA_AZUL, size=10)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
            row += 1
            for act in d["actions"]:
                ws.cell(row=row, column=1).value = f"   → {act}"
                ws.cell(row=row, column=1).font = Font(color=SURA_GRIS_OSC, size=10, italic=True)
                ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
                ws.row_dimensions[row].height = 26
                row += 1
        row += 1  # spacer

    # Apéndice: explicación de la fórmula
    row += 1
    ws.cell(row=row, column=1).value = "📐 Cómo leer este diagnóstico (apéndice metodológico)"
    style_subheader(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    row += 1
    metodo = [
        "El CPL (costo por conversión que reporta la plataforma) es matemáticamente:",
        "   CPL = Spend ÷ Conversiones = CPC × (1 ÷ Tasa de Conversión)",
        "Por lo tanto, un aumento del CPL puede venir de 2 lados:",
        "   1) CPC sube (subasta más cara, pujas más altas, peor Quality Score)",
        "   2) Tasa de Conversión cae (tráfico de menor calidad, landing rota, formulario fallando)",
        "El driver dominante indica la palanca a accionar:",
        "   - Si sube CPC y la tasa conv se mantiene → problema de pujas/competencia, no de funnel",
        "   - Si cae tasa conv y CPC se mantiene → problema de funnel, no de bidding",
        "   - Si suben ambos → problema doble (auditar ambos lados)",
        "Para Meta, en lugar de CPC trabajamos con CPM (Meta cobra por impresión, no por click).",
        "",
        "Severidad: CRÍTICO ≥ +80% CPL · ALTO 40-80% · MEDIO 20-40% · OK ≤ 20%",
    ]
    for line in metodo:
        ws.cell(row=row, column=1).value = line
        ws.cell(row=row, column=1).font = Font(color=SURA_GRIS_OSC, size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        row += 1


def write_combinado(ws, google_agg, meta_agg):
    ws.title = "Combinado"
    for col in "ABCDEFGHIJKLM":
        ws.column_dimensions[col].width = 13
    ws.column_dimensions["A"].width = 16

    ws["A1"] = "Comparativo Google vs Meta · CPA por producto y mes (US$)"
    ws["A1"].font = Font(bold=True, size=14, color=SURA_AZUL)
    ws.merge_cells("A1:M1")

    row = 3
    headers = ["Producto", "Plataforma"] + MONTHS_SHORT + ["Promedio"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i)
        c.value = h
        style_header(c)
    row += 1
    table_start = row
    for prod in PRODUCTS:
        for plat, agg in [("Google", google_agg), ("Meta", meta_agg)]:
            ws.cell(row=row, column=1).value = prod
            ws.cell(row=row, column=1).font = Font(bold=True, color=SURA_AZUL)
            ws.cell(row=row, column=1).border = border
            ws.cell(row=row, column=2).value = plat
            ws.cell(row=row, column=2).border = border
            ws.cell(row=row, column=2).font = Font(color=SURA_GRIS_OSC, italic=True)
            cpas = []
            for j, m in enumerate(MONTHS_ORDER):
                v = agg[prod][m]["cpa"]
                c = ws.cell(row=row, column=3 + j)
                c.value = round(v, 2) if v else 0
                c.number_format = '"$"#,##0.00'
                style_data(c)
                if v > 0:
                    cpas.append(v)
            avg = sum(cpas) / len(cpas) if cpas else 0
            c = ws.cell(row=row, column=7)
            c.value = round(avg, 2)
            c.number_format = '"$"#,##0.00'
            style_total(c)
            row += 1

    # Coste comparison
    row += 2
    ws.cell(row=row, column=1).value = "Inversión total Ene–Abr (US$) por plataforma"
    style_subheader(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    inv_header = ["Producto", "Google", "Meta", "Total", "% Google"]
    for i, h in enumerate(inv_header, start=1):
        c = ws.cell(row=row, column=i)
        c.value = h
        style_header(c)
    row += 1
    inv_start = row
    for prod in PRODUCTS:
        gtot = sum(google_agg[prod][m]["coste"] for m in MONTHS_ORDER)
        mtot = sum(meta_agg[prod][m]["coste"] for m in MONTHS_ORDER)
        tot = gtot + mtot
        ws.cell(row=row, column=1).value = prod
        ws.cell(row=row, column=1).font = Font(bold=True, color=SURA_AZUL)
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2).value = round(gtot, 0)
        ws.cell(row=row, column=2).number_format = '"$"#,##0'
        style_data(ws.cell(row=row, column=2))
        ws.cell(row=row, column=3).value = round(mtot, 0)
        ws.cell(row=row, column=3).number_format = '"$"#,##0'
        style_data(ws.cell(row=row, column=3))
        ws.cell(row=row, column=4).value = round(tot, 0)
        ws.cell(row=row, column=4).number_format = '"$"#,##0'
        style_total(ws.cell(row=row, column=4))
        ws.cell(row=row, column=5).value = (gtot / tot) if tot else 0
        ws.cell(row=row, column=5).number_format = '0%'
        style_data(ws.cell(row=row, column=5))
        row += 1

    # Chart CPA Google vs Meta side by side
    chart = LineChart()
    chart.title = "CPA mensual por producto y plataforma"
    chart.style = 12
    chart.height = 12
    chart.width = 22
    chart.y_axis.title = "CPA (US$)"
    chart.x_axis.title = "Mes"
    data = Reference(ws, min_col=2, max_col=6, min_row=table_start, max_row=row - 1 - 4)
    chart.add_data(data, titles_from_data=False, from_rows=True)
    cats = Reference(ws, min_col=3, max_col=6, min_row=table_start - 1, max_row=table_start - 1)
    chart.set_categories(cats)
    # Custom series labels
    series_labels = []
    for prod in PRODUCTS:
        series_labels.append(f"{prod} Google")
        series_labels.append(f"{prod} Meta")
    for i, ser in enumerate(chart.series):
        if i < len(series_labels):
            ser.tx = None  # rely on auto title from row
    ws.add_chart(chart, "I3")


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def main():
    # Parse Google
    rows_by_product = {}
    for prod, fname in GOOGLE_FILES.items():
        path = GOOGLE_DIR / fname
        if not path.exists():
            print(f"[WARN] missing {path}")
            rows_by_product[prod] = []
            continue
        rows = parse_google_csv(path)
        rows_by_product[prod] = rows
        print(f"[OK] Google {prod:10s}: {len(rows):3d} filas (campañas x mes)")

    # Parse Meta
    if not META_PATH.exists():
        print(f"[ERR] missing {META_PATH}")
        return
    meta_rows = parse_meta_csv(META_PATH)
    print(f"[OK] Meta total: {len(meta_rows)} filas (ad sets x mes)")

    # Aggregate
    google_agg = aggregate_google(rows_by_product)
    meta_agg = aggregate_meta(meta_rows)

    # Quick summary print
    print("\nResumen Google (Coste US$ / Conversiones / CPA):")
    for prod in PRODUCTS:
        for mes in MONTHS_ORDER:
            a = google_agg[prod][mes]
            print(f"  {prod:10s} {mes:8s}  ${a['coste']:>9,.0f}  {a['conv']:>7,.0f}  ${a['cpa']:>5.2f}")
    print("\nResumen Meta (Coste US$ / Conversiones / CPA):")
    for prod in PRODUCTS:
        for mes in MONTHS_ORDER:
            a = meta_agg[prod][mes]
            print(f"  {prod:10s} {mes:8s}  ${a['coste']:>9,.0f}  {a['conv']:>7,.0f}  ${a['cpa']:>5.2f}")

    # Build Excel
    wb = Workbook()
    ws_resumen = wb.active
    write_resumen(ws_resumen, google_agg, meta_agg)
    ws_google = wb.create_sheet("Google Ads")
    write_platform(ws_google, "Google Ads", google_agg, SURA_AZUL)
    ws_meta = wb.create_sheet("Meta Ads")
    write_platform(ws_meta, "Meta Ads", meta_agg, "1877F2")
    ws_comb = wb.create_sheet("Combinado")
    write_combinado(ws_comb, google_agg, meta_agg)

    # CPL Negocio (cruce con Sheet BI)
    bi_leads = load_bi_leads_by_product_month_source()
    bi_data = load_bi_full()
    bi_index = index_bi_by_campaign(bi_data) if bi_data else {}

    if bi_leads:
        ws_cpl = wb.create_sheet("CPL Negocio")
        write_cpl_negocio(ws_cpl, google_agg, meta_agg, bi_leads)

    if bi_index:
        ws_match = wb.create_sheet("Match Campañas")
        write_match_campanas(ws_match, rows_by_product, meta_rows, bi_index)

    if bi_data:
        ws_cac = wb.create_sheet("CAC Real")
        write_cac(ws_cac, google_agg, meta_agg, bi_data)

    # Diagnóstico forense del CPL para CEO
    google_full = aggregate_google_full(rows_by_product)
    meta_full = aggregate_meta_full(meta_rows)
    ws_diag = wb.create_sheet("Diagnóstico CEO")
    write_diagnostico_ceo(ws_diag, google_full, meta_full)

    # Console summary
    print("\nCPL Negocio (Spend / Leads SF) por producto-plataforma totales:")
    for prod in PRODUCTS:
        for plat, agg in [("Google", google_agg), ("Meta", meta_agg)]:
            spend = sum(agg[prod][m]["coste"] for m in MONTHS_ORDER)
            bi_key = "google" if plat == "Google" else "meta"
            leads = sum(bi_leads.get(prod, {}).get(m, {}).get(bi_key, 0) for m in MONTHS_ORDER)
            cpl = spend / leads if leads else 0
            conv = sum(agg[prod][m]["conv"] for m in MONTHS_ORDER)
            cpa = spend / conv if conv else 0
            ratio = leads / conv if conv else 0
            print(f"  {prod:10s} {plat:7s} spend=${spend:>9,.0f} leadsSF={leads:>5,} CPL=${cpl:>5.2f} | CPA=${cpa:>5.2f} ratio={ratio*100:>5.0f}%")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"\n[OK] Excel generado: {OUT_PATH}  ({OUT_PATH.stat().st_size / 1024:.1f} kB)")


if __name__ == "__main__":
    main()
