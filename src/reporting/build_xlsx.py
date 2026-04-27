"""Genera outputs/suratech_data_consolidado.xlsx con 5 hojas:
  Resumen | Detalle Google Ads | Eventos | Discrepancias | Recomendaciones
"""

from __future__ import annotations
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.config import DOCS_DATA, OUTPUTS_DIR

SURA_AZUL  = "00359C"
SURA_AQUA  = "00AEC7"
SURA_GRIS  = "F3F4F6"
ROJO       = "DC2626"
AMARILLO   = "F59E0B"
VERDE      = "10B981"

H1 = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
H2 = Font(bold=True, color=SURA_AZUL, size=11, name="Calibri")
NORMAL = Font(name="Calibri", size=10)
THIN = Side(border_style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_AZUL = PatternFill("solid", fgColor=SURA_AZUL)
FILL_GRIS = PatternFill("solid", fgColor=SURA_GRIS)
FILL_ROJO = PatternFill("solid", fgColor="FEF2F2")
FILL_AMA  = PatternFill("solid", fgColor="FFFBEB")
FILL_VER  = PatternFill("solid", fgColor="ECFDF5")


def _autosize(ws, max_w: int = 60) -> None:
    for col_idx, col in enumerate(ws.columns, 1):
        m = 8
        for c in col:
            v = "" if c.value is None else str(c.value)
            m = max(m, min(len(v) + 2, max_w))
        ws.column_dimensions[get_column_letter(col_idx)].width = m


def _header_row(ws, row: int, headers: list) -> None:
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = H1
        cell.fill = FILL_AZUL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER


def write_resumen(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Resumen")
    ws["A1"] = "SURA Tech Colombia · Informe mensual"
    ws["A1"].font = Font(bold=True, size=18, color=SURA_AZUL)
    ws.merge_cells("A1:F1")
    ws["A2"] = f"Periodo: {data['meta']['periodo']['mes_actual']} · Generado: {data['meta']['generated_at']}"
    ws["A2"].font = Font(italic=True, size=10, color="6B7280")
    ws.merge_cells("A2:F2")

    # KPIs
    ws["A4"] = "KPIs globales"
    ws["A4"].font = H2
    k = data["kpis_globales"]
    rows = [
        ("Inversión total (USD)",       k["inversion_total"]),
        ("  Inversión Google",          k["inversion_google"]),
        ("  Inversión Meta",            k["inversion_meta"]),
        ("  Inversión Bing",            k["inversion_bing"]),
        ("RECIBIDOS (SF) - leads reales", k["leads_recibidos_sf"]),
        ("Total Pauta",                 k["leads_pauta_total"]),
        ("Requeridos (compromiso)",     k["leads_requeridos"]),
        ("Cumplimiento SF/Req",         f"{k['cumpl_sf_vs_req']*100:.1f}%"),
        ("Cumplimiento Pauta/Req",      f"{k['cumpl_pauta_vs_req']*100:.1f}%"),
        ("CPL Negocio promedio (SF)",   k["cpl_negocio_promedio"]),
    ]
    for i, (label, val) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=label).font = NORMAL
        c = ws.cell(row=i, column=2, value=val); c.font = Font(bold=True, size=11)
        c.alignment = Alignment(horizontal="right")

    # Por seguro
    start_row = 12
    ws.cell(row=start_row, column=1, value="Por seguro").font = H2
    headers = ["Seguro","Tipo","Requeridos","RECIBIDOS (SF)","Total Pauta","Cumpl SF/Req","Cumpl Pauta/Req","CPL Negocio (SF)","Inv. Google","Inv. Meta","Inv. Bing","Campañas Google"]
    _header_row(ws, start_row + 1, headers)
    r = start_row + 2
    transcurrido = data["meta"]["periodo"].get("porcentaje_transcurrido", 0.7)
    for s in data["seguros"]:
        c = s["crm"]
        cump = c.get("cumpl_sf_vs_req")
        if cump:
            ratio = cump / transcurrido
            fill = FILL_VER if ratio >= 0.9 else (FILL_AMA if ratio >= 0.6 else FILL_ROJO)
        else:
            fill = FILL_GRIS
        vals = [
            s["nombre"],
            "Principal" if s["es_principal"] else "Extra (solo ads)",
            c.get("requeridos") or "—",
            c.get("recibidos_sf") or "—",
            c.get("total_pauta") or "—",
            f"{cump*100:.1f}%" if cump else "—",
            f"{(c.get('cumpl_pauta_vs_req') or 0)*100:.1f}%" if c.get('cumpl_pauta_vs_req') else "—",
            c.get("cpl_negocio_sf") or "—",
            c.get("consumo_google") or 0,
            c.get("consumo_meta") or 0,
            c.get("consumo_bing") or 0,
            s["google_ads"]["campanias_count"],
        ]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.fill = fill
            cell.border = BORDER
            cell.font = NORMAL
        r += 1
    _autosize(ws)


def write_detalle_ads(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Detalle Google Ads")
    headers = ["Seguro","Campaña","Estado","Tipo","Presupuesto","Impresiones","Clicks","CTR %","CPC","Coste USD","Conversiones","CPA","Estrategia"]
    _header_row(ws, 1, headers)
    r = 2
    for s in data["seguros"]:
        for c in s["google_ads"]["campanias"]:
            row = [
                s["nombre"], c["nombre"], c["estado"], c.get("tipo",""),
                c.get("presupuesto_str") or c.get("presupuesto_diario") or "",
                c["impresiones"], c["clicks"],
                f"{c['ctr_pct']*100:.2f}%" if c["ctr_pct"] else "0%",
                c["cpc"], c["coste"], c["conversiones"], c["cpa"], c.get("estrategia",""),
            ]
            for i, v in enumerate(row, 1):
                cell = ws.cell(row=r, column=i, value=v)
                cell.border = BORDER
                cell.font = NORMAL
                if "limit" in (c["estado"] or "").lower() or "rechaz" in (c["estado"] or "").lower():
                    cell.fill = FILL_ROJO
            r += 1
    _autosize(ws)


def write_discrepancias(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Discrepancias")
    ws["A1"] = "Discrepancias y gaps"
    ws["A1"].font = Font(bold=True, size=14, color=SURA_AZUL)
    d = data.get("discrepancias_y_gaps", {})
    r = 3
    for k, v in d.items():
        ws.cell(row=r, column=1, value=k).font = H2
        ws.cell(row=r+1, column=1, value=v).font = NORMAL
        ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=4)
        ws.cell(row=r+1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r+1].height = 50
        r += 3
    _autosize(ws, max_w=80)


def write_recomendaciones(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Recomendaciones")
    headers = ["#","Prioridad","Seguro","Plataforma","Título","Qué hacer","Por qué","Impacto","Esfuerzo","Plazo"]
    _header_row(ws, 1, headers)
    r = 2
    for i, reco in enumerate(data.get("recomendaciones", []), 1):
        fill = FILL_ROJO if reco["prioridad"] == "critica" else (FILL_AMA if reco["prioridad"] == "alta" else FILL_VER)
        row = [
            i, reco["prioridad"].upper(), reco["seguro"], reco["plataforma"],
            reco["titulo"], reco["que_hacer"], reco["por_que"],
            reco["impacto"], reco["esfuerzo"], reco["plazo"],
        ]
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.fill = fill
            cell.border = BORDER
            cell.font = NORMAL
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 60
        r += 1
    # Anchos custom
    widths = [4, 12, 18, 12, 40, 50, 50, 30, 10, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_evolucion(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Evolución MoM")
    headers = ["Seguro","Mes","Requeridos","RECIBIDOS (SF)","Total Pauta","Cumpl SF/Req","Cumpl Pauta/Req","CPL Neg (SF)","Inv Google","Leads Google","Inv Meta","Leads Meta","Inv Bing","Leads Bing"]
    _header_row(ws, 1, headers)
    r = 2
    for seguro, serie in data.get("evolucion_mom", {}).get("por_seguro", {}).items():
        for p in serie:
            if p.get("vacio"): continue
            row = [seguro, p["mes"], p["requeridos"], p["recibidos_sf"], p["total_pauta"],
                   f"{p['cumpl_sf_vs_req']*100:.1f}%", f"{p['cumpl_pauta_vs_req']*100:.1f}%",
                   p["cpl_negocio_sf"],
                   p["consumo_google"], p["leads_google"],
                   p["consumo_meta"], p["leads_meta"],
                   p["consumo_bing"], p["leads_bing"]]
            for i, v in enumerate(row, 1):
                cell = ws.cell(row=r, column=i, value=v)
                cell.border = BORDER
                cell.font = NORMAL
            r += 1
    _autosize(ws)


def write_cruces(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Cruces")
    headers = ["Seguro","RECIBIDOS (SF)","Total Pauta","Conv. Google Ads","Ratio GA/SF","Ratio Pauta/SF","Leads Google Sheet","Leads Meta Sheet","Leads Bing Sheet","Inv GA Extractor","Inv GA Sheet","Δ Inversión"]
    _header_row(ws, 1, headers)
    r = 2
    for c in data.get("cruces", {}).get("por_seguro", []):
        ratio = c.get("ratio_ga_conv_vs_sf", 0)
        fill = FILL_ROJO if ratio == 0 else (FILL_AMA if ratio < 0.5 else FILL_VER)
        row = [c["seguro"], c["recibidos_sf"], c["total_pauta"], c["google_ads_conv"],
               f"{ratio*100:.0f}%", f"{c.get('ratio_pauta_vs_sf',0)*100:.0f}%",
               c["leads_google_sheet"], c["leads_meta_sheet"], c["leads_bing_sheet"],
               c["inversion_ga_extractor"], c["inversion_ga_sheet"], c["delta_inversion"]]
        for i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.fill = fill; cell.border = BORDER; cell.font = NORMAL
        r += 1
    _autosize(ws)


def write_optimizaciones(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Optimizaciones")
    cats = [("escalar","ESCALAR (CPA bajo + headroom)", FILL_VER),
            ("pausar", "PAUSAR (CPA alto)",             FILL_ROJO),
            ("destrabar","DESTRABAR (rechazos/políticas)", FILL_AMA),
            ("quality_issues","QUALITY ISSUES",          FILL_AMA)]
    r = 1
    for cat_key, cat_label, fill in cats:
        ws.cell(row=r, column=1, value=cat_label).font = Font(bold=True, color=SURA_AZUL, size=12)
        r += 1
        headers = ["Seguro","Campaña","Subtipo","Budget/día","Coste","Conversiones","CPA","% Lost Budget","% Lost Rank","Estado"]
        _header_row(ws, r, headers)
        r += 1
        for c in data.get("optimizaciones", {}).get(cat_key, []):
            row = [c["seguro"], c["nombre"], c["subtipo"], c["presupuesto_diario"],
                   c["coste"], int(c["conversiones"]), c["cpa"],
                   f"{c['cuota_perdida_budget']*100:.0f}%",
                   f"{c['cuota_perdida_ranking']*100:.0f}%",
                   c["estado"]]
            for i, v in enumerate(row, 1):
                cell = ws.cell(row=r, column=i, value=v)
                cell.fill = fill; cell.border = BORDER; cell.font = NORMAL
            r += 1
        r += 2
    _autosize(ws, max_w=50)


def run() -> int:
    data = json.loads(DOCS_DATA.read_text(encoding="utf-8"))
    wb = Workbook()
    wb.remove(wb.active)
    write_resumen(wb, data)
    write_detalle_ads(wb, data)
    write_evolucion(wb, data)
    write_cruces(wb, data)
    write_optimizaciones(wb, data)
    write_discrepancias(wb, data)
    write_recomendaciones(wb, data)

    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    out = OUTPUTS_DIR / "suratech_data_consolidado.xlsx"
    wb.save(out)
    print(f"[xlsx] -> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
