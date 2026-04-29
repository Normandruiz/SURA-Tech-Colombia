"""Genera outputs/suratech_propuesta_distribucion_mayo2026.xlsx con la propuesta
de redistribucion de presupuesto por seguro x medio x campana.

Constraints:
  - Tope por seguro (provisto por user):
      Auto:      $52.421
      Moto:      $32.642
      Viajes:    $16.900
      Arriendo:  $31.080
  - Bing: dejar en su plan del Flow (no se toca)
  - 70-75% del presupuesto debe ir a Google (mejor calidad de leads)
"""

from __future__ import annotations
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.config import OUTPUTS_DIR, ROOT


SURA_AZUL = "0033A0"
SURA_AQUA = "00AEC7"
GRIS_FONDO = "F0F0F0"
ROJO   = "DC2626"
AMARILLO = "F59E0B"
VERDE  = "10B981"

BORDER = Border(
    left=Side("thin", color="DDDDDD"),
    right=Side("thin", color="DDDDDD"),
    top=Side("thin", color="DDDDDD"),
    bottom=Side("thin", color="DDDDDD"),
)
H1 = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
H_AZUL = PatternFill("solid", fgColor=SURA_AZUL)
NORMAL = Font(name="Calibri", size=10)
BOLD = Font(name="Calibri", size=10, bold=True)


# ============== INPUT: DATA REAL DEL 1-27 ABRIL ==============

BUDGETS = {
    "Auto":      52421,
    "Moto":      32642,
    "Viajes":    16900,
    "Arriendo":  31080,
}

# Spend / leads / CPA del 1-27 abr 2026 (calculado del Sheet)
ACTUAL = {
    "Auto":     {"google": (26132, 4337, 6.03), "meta": (10244, 2544, 4.03), "bing": 1120},
    "Moto":     {"google": (15555, 1117, 13.92), "meta": (6629, 1192, 5.56), "bing": 753},
    "Arriendo": {"google": (28902, 1870, 15.45), "meta": (8083, 1508, 5.36), "bing": 1167},
    "Viajes":   {"google": (6762, 204, 33.01), "meta": (3186, 83, 38.38), "bing": 178},
}

def _load_google_camps_from_dataset():
    """Lee las campanias Google del dataset paid_media (ya tiene clicks/conv reales)."""
    p = ROOT / "docs" / "assets" / "data_paid_media.json"
    if not p.exists():
        return {}
    d = json.loads(p.read_text(encoding="utf-8"))
    out = {}
    map_seguro = {"Autos": "Auto", "Motos": "Moto", "Arrendamiento": "Arriendo", "Viajes": "Viajes"}
    for nombre_full, b in d.get("seguros", {}).items():
        seguro = map_seguro.get(nombre_full)
        if not seguro:
            continue
        out[seguro] = []
        for c in b.get("campanias", []):
            clicks = c.get("clicks") or 0
            conv = c.get("conversiones") or 0
            tasa = (conv / clicks) if clicks else 0
            out[seguro].append({
                "nombre":          c["nombre"],
                "cpa":             c.get("cpa") or 0,
                "lost_budget":     c.get("cuota_perdida_budget") or 0,
                "cuota_perdida_ranking": c.get("cuota_perdida_ranking") or 0,
                "budget_d_actual": c.get("presupuesto_diario") or 0,
                "spend":           c.get("coste") or 0,
                "conv":            conv,
                "clicks":          clicks,
                "tasa_conv":       tasa,
                "ctr":             c.get("ctr_pct") or 0,
                "flags":           c.get("flags") or [],
            })
    return out


GOOGLE_CAMPS = _load_google_camps_from_dataset()


# ============== ALGORITMO DE PROPUESTA ==============

# Reglas:
#  - Cumplir 70-75% Google del total por seguro
#  - Bing fijo (lo del Flow)
#  - Por campania Google: clasificar en escalar / mantener / recortar segun CPA + lost_budget
#  - Asignar el budget Google priorizando las que tienen mejor CPA y mas lost_budget

def clasificar(c):
    """Devuelve (accion, factor) considerando CPA + lost_budget + tasa de conversion.

    Tasa de conversion (conv/clicks) cambia el peso de las decisiones:
    - tasa alta (>8%) refuerza el escalar
    - tasa media (4-8%) accion estandar
    - tasa baja (<4%) penaliza el escalado y refuerza el recorte
    """
    cpa = c["cpa"]; lb = c["lost_budget"]; tc = c.get("tasa_conv", 0)
    # Buckets de tasa de conversion
    if   tc >= 0.08: tier = "alta"
    elif tc >= 0.04: tier = "media"
    elif tc > 0:     tier = "baja"
    else:            tier = "sin"

    # CPA bajo ($<=7)
    if cpa > 0 and cpa <= 7:
        if lb >= 0.50 and tier in ("alta", "media"):
            return "ESCALAR FUERTE", 1.80
        if lb >= 0.30:
            return "ESCALAR", 1.40
        if lb >= 0.10:
            return "ESCALAR LEVE", 1.20
        return "MANTENER", 1.0

    # CPA medio ($7-10)
    if cpa > 0 and cpa <= 10:
        if lb >= 0.30 and tier == "alta":
            return "ESCALAR", 1.40
        if lb >= 0.10 and tier in ("alta", "media"):
            return "ESCALAR LEVE", 1.20
        return "MANTENER", 1.0

    # CPA medio-alto ($10-15)
    if cpa > 0 and cpa <= 15:
        if lb >= 0.50 and tier == "alta":
            return "ESCALAR MODERADO", 1.15  # solo si tasa conv es alta
        if tier == "baja":
            return "RECORTAR", 0.65
        return "MANTENER", 1.0

    # CPA alto ($15-25)
    if cpa > 0 and cpa <= 25:
        return "RECORTAR", 0.65

    # CPA insostenible (>$25)
    if cpa > 25:
        return "RECORTAR FUERTE", 0.40

    return "MANTENER", 1.0


def proponer(seguro: str):
    actual = ACTUAL[seguro]
    budget_total = BUDGETS[seguro]
    bing = actual["bing"]
    spend_g, leads_g, cpa_g = actual["google"]
    spend_m, leads_m, cpa_m = actual["meta"]

    # Target Google = 72% del budget total (rango 70-75)
    google_target = round(budget_total * 0.72)
    meta_target   = budget_total - google_target - bing

    # Distribucion entre campanias Google
    camps = GOOGLE_CAMPS[seguro]
    rows = []
    # Calculo inicial usando el factor sugerido por la regla
    sum_propuesto = 0
    for c in camps:
        accion, factor = clasificar(c)
        nuevo = round(c["spend"] * factor)
        sum_propuesto += nuevo
        rows.append({**c, "accion": accion, "factor_inicial": factor, "spend_propuesto_pre": nuevo})

    # Reescalar para cuadrar con google_target (proporcional)
    if sum_propuesto > 0:
        ratio = google_target / sum_propuesto
        for r in rows:
            r["spend_propuesto"] = round(r["spend_propuesto_pre"] * ratio)
            r["budget_d_propuesto"] = round(r["spend_propuesto"] / 27, 0)  # 27 dias
            r["delta"] = r["spend_propuesto"] - r["spend"]
            r["delta_pct"] = (r["delta"] / r["spend"] * 100) if r["spend"] else 0

    return {
        "seguro":          seguro,
        "budget_total":    budget_total,
        "actual": {
            "google_spend":  spend_g,
            "google_leads":  leads_g,
            "google_cpa":    cpa_g,
            "meta_spend":    spend_m,
            "meta_leads":    leads_m,
            "meta_cpa":      cpa_m,
            "bing_spend":    bing,
            "total":         spend_g + spend_m + bing,
        },
        "propuesta": {
            "google_total":  sum(r["spend_propuesto"] for r in rows),
            "meta_total":    meta_target,
            "bing":          bing,
            "google_pct":    sum(r["spend_propuesto"] for r in rows) / budget_total * 100,
            "meta_pct":      meta_target / budget_total * 100,
            "bing_pct":      bing / budget_total * 100,
        },
        "campanias_google": rows,
    }


# ============== EXCEL WRITER ==============

def header(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = H1; c.fill = H_AZUL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def write_resumen(wb: Workbook, props: list[dict]) -> None:
    ws = wb.create_sheet("Resumen")
    ws["A1"] = "SURA Tech Colombia · Propuesta distribución Mayo 2026"
    ws["A1"].font = Font(bold=True, size=16, color=SURA_AZUL)
    ws.merge_cells("A1:I1")
    ws["A2"] = "Beyond Media Agency · 70-75% Google · respetando topes por seguro · Bing fijo en plan"
    ws["A2"].font = Font(italic=True, size=10, color="768692")
    ws.merge_cells("A2:I2")
    r = 4
    headers = ["Seguro", "Budget total", "Google ahora", "Google propuesto", "Δ Google",
               "Meta ahora", "Meta propuesto", "Δ Meta", "Bing fijo"]
    widths = [16, 13, 14, 16, 13, 13, 14, 13, 11]
    header(ws, r, headers, widths)
    r += 1
    tot = {"budget":0, "g_now":0, "g_new":0, "m_now":0, "m_new":0, "b":0}
    for p in props:
        a = p["actual"]; pr = p["propuesta"]
        row = [p["seguro"], p["budget_total"],
               a["google_spend"], pr["google_total"], pr["google_total"] - a["google_spend"],
               a["meta_spend"], pr["meta_total"], pr["meta_total"] - a["meta_spend"],
               pr["bing"]]
        for i, v in enumerate(row, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDER; c.font = NORMAL
            if i == 1: c.font = BOLD
            if i in (5, 8):
                if isinstance(v, (int, float)):
                    if v > 0:
                        c.font = Font(name="Calibri", size=10, bold=True, color=VERDE)
                    elif v < 0:
                        c.font = Font(name="Calibri", size=10, bold=True, color=ROJO)
            if i in (2, 3, 4, 6, 7, 9):
                c.number_format = '"$"#,##0'
            if i == 5 or i == 8:
                c.number_format = '"$"#,##0;[RED]"-$"#,##0'
        tot["budget"] += p["budget_total"]
        tot["g_now"]  += a["google_spend"]
        tot["g_new"]  += pr["google_total"]
        tot["m_now"]  += a["meta_spend"]
        tot["m_new"]  += pr["meta_total"]
        tot["b"]      += pr["bing"]
        r += 1
    # Fila TOTAL
    ws.cell(row=r, column=1, value="TOTAL").font = BOLD
    ws.cell(row=r, column=2, value=tot["budget"])
    ws.cell(row=r, column=3, value=tot["g_now"])
    ws.cell(row=r, column=4, value=tot["g_new"])
    ws.cell(row=r, column=5, value=tot["g_new"] - tot["g_now"])
    ws.cell(row=r, column=6, value=tot["m_now"])
    ws.cell(row=r, column=7, value=tot["m_new"])
    ws.cell(row=r, column=8, value=tot["m_new"] - tot["m_now"])
    ws.cell(row=r, column=9, value=tot["b"])
    for col in range(1, 10):
        c = ws.cell(row=r, column=col)
        c.font = BOLD; c.border = BORDER
        c.fill = PatternFill("solid", fgColor=GRIS_FONDO)
        if col in (2, 3, 4, 5, 6, 7, 8, 9):
            c.number_format = '"$"#,##0'

    # Bloque 2: % Google / Meta / Bing por seguro (validacion del 70-75)
    r += 3
    ws.cell(row=r, column=1, value="% del total por canal").font = Font(bold=True, color=SURA_AZUL, size=12)
    r += 1
    header(ws, r, ["Seguro", "% Google", "% Meta", "% Bing", "Suma", "Cumple 70-75% Google?"],
           [16, 12, 12, 12, 12, 22])
    r += 1
    for p in props:
        pr = p["propuesta"]
        suma = pr["google_pct"] + pr["meta_pct"] + pr["bing_pct"]
        cumple = "✓ Sí" if 68 <= pr["google_pct"] <= 77 else "⚠ Revisar"
        row = [p["seguro"],
               f"{pr['google_pct']:.1f}%",
               f"{pr['meta_pct']:.1f}%",
               f"{pr['bing_pct']:.1f}%",
               f"{suma:.1f}%",
               cumple]
        for i, v in enumerate(row, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDER; c.font = NORMAL
            if i == 1: c.font = BOLD
            if i == 6 and "⚠" in str(v):
                c.font = Font(color=AMARILLO, bold=True)
            elif i == 6:
                c.font = Font(color=VERDE, bold=True)
        r += 1


def _justificar_google(c, accion):
    cpa = c["cpa"]; lb = c["lost_budget"]
    lr = c.get("cuota_perdida_ranking", 0)
    conv = c["conv"]; spend = c["spend"]; flags = c.get("flags", [])
    tc = c.get("tasa_conv", 0); clicks = c.get("clicks", 0)
    tc_str = f"{tc*100:.1f}%" if tc > 0 else "—"
    if   tc >= 0.08: tc_eval = "alta"
    elif tc >= 0.04: tc_eval = "media"
    elif tc > 0:     tc_eval = "baja"
    else:            tc_eval = "sin"

    if "ESCALAR FUERTE" in accion:
        return (f"CPA ${cpa:.2f} + tasa de conversión {tc_str} ({tc_eval}) + pierde {lb*100:.0f}% por presupuesto. "
                f"Combo ideal: convierte bien y le falta plata. Subir budget convierte impresiones perdidas en conversiones al mismo CPA.")

    if accion == "ESCALAR" or accion == "ESCALAR LEVE" or accion == "ESCALAR MODERADO":
        partes = [f"CPA ${cpa:.2f}", f"tasa conv {tc_str} ({tc_eval})"]
        if conv >= 200: partes.append(f"{int(conv):,} conversiones validan el formato")
        if lb >= 0.10: partes.append(f"{lb*100:.0f}% lost budget = headroom")
        return ", ".join(partes) + ". Subir budget para capitalizar la eficiencia."

    if "RECORTAR FUERTE" in accion:
        return (f"CPA ${cpa:.2f} insostenible. Tasa conv {tc_str} ({tc_eval}) sobre {int(clicks):,} clicks. "
                f"${spend:,.0f} invertidos generaron solo {int(conv)} conversiones. Bajar budget y revisar keywords/landing antes de seguir gastando.")

    if "RECORTAR" in accion:
        if tc_eval == "baja":
            return (f"Tasa de conversión {tc_str} ({tc_eval}) sobre {int(clicks):,} clicks → el funnel no convierte bien. "
                    f"CPA ${cpa:.2f}. Recortar y revisar landing/oferta.")
        if "rejected" in flags:
            return f"CPA ${cpa:.2f} + tasa conv {tc_str} + anuncios rechazados. Recortar mientras se corrigen los assets."
        if lb < 0.10 and cpa > 12:
            return f"CPA ${cpa:.2f} alto sin headroom (lost {lb*100:.0f}%) y tasa conv {tc_str}. El problema no es budget, es formato/intent."
        return f"CPA ${cpa:.2f} fuera de rango con tasa conv {tc_str}. Recortar y mover budget a campañas más eficientes."

    if "MANTENER" in accion:
        if cpa > 0 and cpa < 12 and lb < 0.10:
            return f"CPA ${cpa:.2f} + tasa conv {tc_str} aceptables, la campaña ya consume su budget (lost {lb*100:.0f}%). Sin headroom para escalar."
        if tc_eval == "baja":
            return f"CPA ${cpa:.2f} pero tasa conv baja ({tc_str}). Mantener para no romper aprendizaje, monitorear performance."
        return f"CPA ${cpa:.2f}, tasa conv {tc_str}. Monitorear y mantener estable."

    return ""


def write_campanias_google(wb: Workbook, props: list[dict]) -> None:
    ws = wb.create_sheet("Detalle Google")
    ws["A1"] = "Detalle por campaña Google Ads"
    ws["A1"].font = Font(bold=True, size=14, color=SURA_AZUL)
    ws.merge_cells("A1:N1")
    r = 3
    headers = ["Seguro", "Campaña", "CPA actual", "Tasa conv.", "% Lost Budget",
               "Budget/día actual", "Spend actual", "Clicks",
               "Conversiones", "Acción", "Spend propuesto", "Δ Spend", "Δ %",
               "Justificación"]
    widths = [13, 42, 11, 11, 12, 13, 12, 11, 12, 19, 14, 11, 8, 75]
    header(ws, r, headers, widths)
    r += 1
    for p in props:
        for c in p["campanias_google"]:
            justif = _justificar_google(c, c["accion"])
            row = [p["seguro"], c["nombre"], c["cpa"],
                   c.get("tasa_conv", 0),
                   c["lost_budget"],
                   c["budget_d_actual"], c["spend"],
                   int(c.get("clicks", 0)),
                   c["conv"],
                   c["accion"], c["spend_propuesto"], c["delta"], c["delta_pct"]/100,
                   justif]
            for i, v in enumerate(row, 1):
                cell = ws.cell(row=r, column=i, value=v)
                cell.border = BORDER; cell.font = NORMAL
                cell.alignment = Alignment(vertical="top", wrap_text=(i == 14))
                if i == 1: cell.font = BOLD
                if i == 3: cell.number_format = '"$"#,##0.00'
                if i == 4: cell.number_format = "0.0%"  # tasa conv
                if i == 5: cell.number_format = "0.0%"  # lost budget
                if i in (6, 7, 11): cell.number_format = '"$"#,##0'
                if i == 8: cell.number_format = "#,##0"
                if i == 9: cell.number_format = "#,##0.0"
                if i == 12: cell.number_format = '"$"#,##0;[RED]"-$"#,##0'
                if i == 13: cell.number_format = '+0.0%;[RED]-0.0%'
                # color de tasa conv: verde si alta, rojo si baja
                if i == 4 and isinstance(v, (int, float)):
                    if v >= 0.08:
                        cell.font = Font(name="Calibri", size=10, color=VERDE, bold=True)
                    elif v < 0.04 and v > 0:
                        cell.font = Font(name="Calibri", size=10, color=ROJO, bold=True)
                if i == 10:
                    if "RECORTAR FUERTE" in str(v):
                        cell.fill = PatternFill("solid", fgColor="FEE2E2"); cell.font = Font(name="Calibri", size=10, bold=True, color=ROJO)
                    elif "RECORTAR" in str(v):
                        cell.fill = PatternFill("solid", fgColor="FEF3C7"); cell.font = Font(name="Calibri", size=10, bold=True, color=AMARILLO)
                    elif "ESCALAR FUERTE" in str(v):
                        cell.fill = PatternFill("solid", fgColor="D1FAE5"); cell.font = Font(name="Calibri", size=10, bold=True, color=VERDE)
                    elif "ESCALAR" in str(v):
                        cell.fill = PatternFill("solid", fgColor="ECFDF5"); cell.font = Font(name="Calibri", size=10, bold=True, color=VERDE)
                    elif "MANTENER" in str(v):
                        cell.fill = PatternFill("solid", fgColor=GRIS_FONDO); cell.font = NORMAL
            ws.row_dimensions[r].height = 60
            r += 1
        r += 1


def _accion_meta(cpr):
    """Clasifica un ad set Meta segun su CPR y devuelve (accion, factor, color)."""
    if cpr <= 0:        return ("SIN DATOS",       1.00, "FFFFFF")
    if cpr < 2:         return ("ESCALAR FUERTE",  1.80, "D1FAE5")
    if cpr < 5:         return ("ESCALAR LEVE",    1.30, "ECFDF5")
    if cpr < 10:        return ("MANTENER",        1.00, "F0F0F0")
    if cpr < 20:        return ("RECORTAR",        0.65, "FEF3C7")
    return                    ("RECORTAR FUERTE",  0.40, "FEE2E2")


def _justificar_meta(d):
    cpr = d["cpr"]; res = d["res"]; spend = d["spend"]
    accion = d["accion"]; indic = (d.get("indic") or "")
    is_msg = "messaging_conversation" in indic or "MENSAJES" in d["name"].upper()
    is_lead = "leadgen" in indic
    tipo = "conversaciones WhatsApp" if is_msg else ("leads (form)" if is_lead else "conversiones web")

    if "ESCALAR FUERTE" in accion:
        return (f"CPR ${cpr:.2f} excepcional ({int(res):,} {tipo}). "
                f"Es un outlier eficiente del MCC - vale la pena duplicar inversión y dejar que escale.")

    if "ESCALAR" in accion:
        return (f"CPR ${cpr:.2f} bueno con {int(res):,} {tipo} validados. "
                f"Subir budget para ganar volumen sin tocar la eficiencia actual.")

    if "RECORTAR FUERTE" in accion:
        return (f"CPR ${cpr:.2f} muy alto. ${spend:,.0f} generaron solo {int(res)} {tipo}. "
                f"Pausar/recortar y reformular antes de seguir gastando.")

    if "RECORTAR" in accion:
        return (f"CPR ${cpr:.2f} arriba del rango eficiente del producto. "
                f"Bajar budget y mover el spend a los ad sets que mejor convierten.")

    if "MANTENER" in accion:
        return (f"CPR ${cpr:.2f} aceptable. {int(res):,} {tipo}. "
                f"Mantener para no romper aprendizaje del algoritmo.")

    return ""


def _load_meta_adsets():
    """Lee CSV exportado del Ads Manager y devuelve lista con seguro inferido."""
    import csv as csvm
    csv_p = ROOT / "raw" / "meta" / "meta_adsets_abr_2026.csv"
    out = []
    if not csv_p.exists():
        return out
    with open(csv_p, encoding="utf-8") as f:
        for r in csvm.DictReader(f):
            name = r["Nombre del conjunto de anuncios"]
            n_up = name.upper()
            if   "AUTO" in n_up:                       seguro = "Auto"
            elif "MOTO" in n_up:                       seguro = "Moto"
            elif "ARRIENDO" in n_up or "AD " in n_up:  seguro = "Arriendo"
            elif "VIAJE" in n_up:                      seguro = "Viajes"
            else:                                       seguro = "OTRO"
            out.append({
                "seguro": seguro,
                "name":   name,
                "res":    float(r["Resultados"] or 0),
                "cpr":    float(r["Costo por resultados"] or 0),
                "spend":  float(r["Importe gastado (USD)"] or 0),
                "indic":  r["Indicador de resultado"],
                "estado": r["Entrega del conjunto de anuncios"],
            })
    return out


def _proponer_meta_por_seguro(detail_meta, props):
    """Asigna el budget Meta de cada seguro entre sus ad sets segun la regla de
    clasificacion + reescala para que la suma cuadre con el target Meta del seguro.

    Devuelve la lista detail_meta con campos extra:
      accion, factor_inicial, spend_propuesto, delta, delta_pct, color
    """
    target_meta = {p["seguro"]: p["propuesta"]["meta_total"] for p in props}

    # Agrupar por seguro
    por_seguro = {}
    for d in detail_meta:
        por_seguro.setdefault(d["seguro"], []).append(d)

    for seguro, ads in por_seguro.items():
        if seguro not in target_meta:
            for d in ads:
                d["accion"], d["factor_inicial"], d["color"] = _accion_meta(d["cpr"])
                d["spend_propuesto"] = d["spend"]
                d["delta"] = 0
                d["delta_pct"] = 0
            continue
        target = target_meta[seguro]

        # Calculo inicial
        sum_pre = 0.0
        for d in ads:
            act, fac, color = _accion_meta(d["cpr"])
            d["accion"] = act; d["factor_inicial"] = fac; d["color"] = color
            d["spend_propuesto_pre"] = d["spend"] * fac
            sum_pre += d["spend_propuesto_pre"]

        # Reescalar a target
        ratio = (target / sum_pre) if sum_pre > 0 else 1.0
        for d in ads:
            d["spend_propuesto"] = round(d["spend_propuesto_pre"] * ratio)
            d["delta"] = d["spend_propuesto"] - d["spend"]
            d["delta_pct"] = (d["delta"] / d["spend"] * 100) if d["spend"] > 0 else 0

    return detail_meta


def write_meta(wb: Workbook, props: list[dict]) -> None:
    """Hoja Meta a nivel ad set con propuesta de presupuesto por ad set."""
    detail_meta = _load_meta_adsets()
    detail_meta = _proponer_meta_por_seguro(detail_meta, props)

    ws = wb.create_sheet("Detalle Meta")
    ws["A1"] = "Detalle Meta — ad sets reales del 1-27 abril 2026"
    ws["A1"].font = Font(bold=True, size=14, color=SURA_AZUL)
    ws.merge_cells("A1:K1")
    n_relevantes = sum(1 for d in detail_meta if d["seguro"] in ("Auto","Moto","Arriendo","Viajes"))
    ws["A2"] = f"Total relevante: {n_relevantes} ad sets en los 4 productos · Fuente: export del Ads Manager"
    ws["A2"].font = Font(italic=True, size=10, color="768692")
    ws.merge_cells("A2:K2")

    r = 4
    headers = ["Seguro", "Conjunto de anuncios", "Resultados", "CPR (USD)",
               "Spend actual", "Indicador",
               "Acción", "Spend propuesto", "Δ Spend", "Δ %", "Estado",
               "Justificación"]
    widths = [13, 44, 11, 11, 12, 38, 18, 14, 12, 9, 11, 70]
    header(ws, r, headers, widths)
    r += 1

    detail_meta.sort(key=lambda d: (d["seguro"], d["cpr"]))
    for d in detail_meta:
        if d["seguro"] in ("OTRO", "Salud Animal", "Salud para Dos"):
            continue
        justif = _justificar_meta(d)
        row = [d["seguro"], d["name"], int(d["res"]), d["cpr"],
               d["spend"], d["indic"][:60],
               d["accion"], d["spend_propuesto"], d["delta"], d["delta_pct"]/100,
               d["estado"], justif]
        for i, v in enumerate(row, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDER; c.font = NORMAL
            c.alignment = Alignment(vertical="top", wrap_text=(i == 12))
            if i == 1: c.font = BOLD
            if i == 3: c.number_format = "#,##0"
            if i == 4: c.number_format = '"$"#,##0.00'
            if i == 5 or i == 8: c.number_format = '"$"#,##0'
            if i == 9: c.number_format = '"$"#,##0;[RED]"-$"#,##0'
            if i == 10: c.number_format = '+0.0%;[RED]-0.0%'
            if i == 7:
                c.fill = PatternFill("solid", fgColor=d["color"])
                if "ESCALAR FUERTE" in str(v):
                    c.font = Font(name="Calibri", size=10, bold=True, color=VERDE)
                elif "ESCALAR" in str(v):
                    c.font = Font(name="Calibri", size=10, bold=True, color=VERDE)
                elif "RECORTAR FUERTE" in str(v):
                    c.font = Font(name="Calibri", size=10, bold=True, color=ROJO)
                elif "RECORTAR" in str(v):
                    c.font = Font(name="Calibri", size=10, bold=True, color=AMARILLO)
                else:
                    c.font = Font(name="Calibri", size=10, color="53565A")
            if i == 10:
                if isinstance(v, (int, float)):
                    if v > 0:
                        c.font = Font(name="Calibri", size=10, color=VERDE, bold=True)
                    elif v < 0:
                        c.font = Font(name="Calibri", size=10, color=ROJO, bold=True)
        ws.row_dimensions[r].height = 50
        r += 1

    # Total por seguro al final
    r += 2
    ws.cell(row=r, column=1, value="Total Meta por seguro").font = Font(bold=True, color=SURA_AZUL, size=12)
    r += 1
    headers2 = ["Seguro", "Ad sets", "Spend actual", "Resultados", "CPR", "Spend propuesto", "Δ Spend"]
    widths2 = [13, 8, 14, 12, 11, 16, 13]
    header(ws, r, headers2, widths2)
    r += 1
    for p in props:
        sel = [d for d in detail_meta if d["seguro"] == p["seguro"]]
        spend = sum(d["spend"] for d in sel)
        prop  = sum(d["spend_propuesto"] for d in sel)
        res   = sum(d["res"] for d in sel)
        cpr   = (spend/res) if res else 0
        row = [p["seguro"], len(sel), spend, int(res), cpr, prop, prop - spend]
        for i, v in enumerate(row, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDER; c.font = NORMAL
            if i == 1: c.font = BOLD
            if i == 2: c.number_format = "0"
            if i == 3 or i == 6: c.number_format = '"$"#,##0'
            if i == 4: c.number_format = "#,##0"
            if i == 5: c.number_format = '"$"#,##0.00'
            if i == 7: c.number_format = '"$"#,##0;[RED]"-$"#,##0'
        r += 1


def write_metodologia(wb: Workbook) -> None:
    ws = wb.create_sheet("Metodología")
    ws["A1"] = "Metodología de la propuesta"
    ws["A1"].font = Font(bold=True, size=14, color=SURA_AZUL)
    ws.merge_cells("A1:B1")
    rows = [
        ("Datos base", "Performance real del 1-27 abril 2026 cruzado con Sheet Leads CRM"),
        ("Constraint 1", "Tope por seguro provisto por el cliente (Auto $52.421, Moto $32.642, Arriendo $31.080, Viajes $16.900)"),
        ("Constraint 2", "70-75% del presupuesto en Google (mejor calidad de leads SF según cliente)"),
        ("Constraint 3", "Bing se mantiene en presupuesto del Flow"),
        ("", ""),
        ("Reglas de clasificación de campañas Google", ""),
        ("Variables consideradas", "CPA + % Lost Budget + Tasa de conversión (conv/clicks)"),
        ("Tasa de conversión", "Buckets: alta ≥ 8% · media 4-8% · baja < 4%"),
        ("ESCALAR FUERTE (×1.80)", "CPA ≤ $7 + Lost Budget ≥ 50% + tasa conv alta o media — palanca #1"),
        ("ESCALAR (×1.40)", "CPA ≤ $10 + Lost Budget ≥ 30%"),
        ("ESCALAR LEVE (×1.20)", "CPA ≤ $10 + Lost Budget ≥ 10% + tasa conv alta o media"),
        ("ESCALAR MODERADO (×1.15)", "CPA $10-15 + Lost Budget ≥ 50% + tasa conv alta (solo si convierte bien)"),
        ("RECORTAR (×0.65)", "CPA $15-25, o tasa conv baja con CPA > $10"),
        ("RECORTAR FUERTE (×0.40)", "CPA > $25"),
        ("MANTENER (×1.00)", "Resto - se monitorea sin tocar"),
        ("", ""),
        ("Ajuste fino", "Después del factor inicial, los presupuestos de Google se reescalan proporcionalmente para alcanzar exactamente el 72% del total por seguro."),
        ("", ""),
        ("Limitaciones honestas", ""),
        ("Meta", "Detalle por ad set requiere CSV exportado del Ads Manager (Reports → Export Reports). Esta propuesta optimiza Meta a nivel seguro."),
        ("Bing", "No se modificó (sin acceso desde el extractor)."),
        ("Estacionalidad", "Asume ritmo similar a abril. Si hay cambios estacionales, ajustar."),
    ]
    r = 3
    for k, v in rows:
        a = ws.cell(row=r, column=1, value=k)
        b = ws.cell(row=r, column=2, value=v)
        a.font = Font(bold=True, color=SURA_AZUL, size=10) if k and not v else BOLD if k else NORMAL
        b.font = NORMAL
        b.alignment = Alignment(wrap_text=True)
        if k and not v:
            a.font = Font(bold=True, color=SURA_AZUL, size=11)
        r += 1
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90


def run() -> int:
    props = [proponer(s) for s in ["Auto", "Moto", "Arriendo", "Viajes"]]

    wb = Workbook()
    wb.remove(wb.active)
    write_resumen(wb, props)
    write_campanias_google(wb, props)
    write_meta(wb, props)
    write_metodologia(wb)

    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    out = OUTPUTS_DIR / "suratech_propuesta_distribucion_mayo2026.xlsx"
    wb.save(out)
    print(f"[propuesta] -> {out}")

    # Tambien dump JSON
    out_json = OUTPUTS_DIR / "suratech_propuesta_distribucion_mayo2026.json"
    out_json.write_text(json.dumps(props, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[propuesta json] -> {out_json}")

    # Stdout summary
    for p in props:
        pr = p["propuesta"]; a = p["actual"]
        print(f"\n=== {p['seguro']} (budget ${p['budget_total']:,}) ===")
        print(f"  Actual:    Google ${a['google_spend']:,} ({a['google_spend']/p['budget_total']*100:.0f}%) / Meta ${a['meta_spend']:,} / Bing ${a['bing_spend']:,}")
        print(f"  Propuesto: Google ${pr['google_total']:,} ({pr['google_pct']:.0f}%) / Meta ${pr['meta_total']:,} ({pr['meta_pct']:.0f}%) / Bing ${pr['bing']:,} ({pr['bing_pct']:.0f}%)")
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
